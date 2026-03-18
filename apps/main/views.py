from django.contrib.auth.decorators import login_required, permission_required
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Sum, Count, Q
from apps.clients.models import Client
from apps.rental.models import RentalOrder, OrderItem, Payment, ReturnDocument
from apps.inventory.models import Product, Category
from datetime import datetime, timedelta
from django.utils import timezone
import json, os
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.http import JsonResponse, FileResponse, Http404
from apps.rental.utils import get_order_groups_for_client, calculate_order_debt
from config import settings
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User, Group, Permission
from .decorators import admin_required, manager_required, cashier_required
from django.contrib.auth import login, authenticate
from apps.main.models import UserProfile
from apps.main.pagination import paginate


def get_tenant_owner(user):
    """
    Возвращает суперпользователя-владельца тенанта для данного user.
    Если user сам суперпользователь — возвращает себя.
    Если user — сотрудник — возвращает owner из его UserProfile.
    """
    if user.is_superuser:
        return user
    try:
        owner = user.profile.owner
        return owner if owner else user
    except Exception:
        return user


def register_view(request):
    """Страница регистрации — только для новых директоров компаний"""
    if request.user.is_authenticated:
        return redirect('main:dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        password_confirm = request.POST.get('password_confirm')
        email = request.POST.get('email')
        phone = request.POST.get('phone', '').strip()

        # Валидация
        if not username or not password:
            messages.error(request, 'Заполните все поля')
            return render(request, 'registration/register.html')

        if password != password_confirm:
            messages.error(request, 'Пароли не совпадают')
            return render(request, 'registration/register.html')

        if User.objects.filter(username=username).exists():
            messages.error(request, 'Пользователь с таким именем уже существует')
            return render(request, 'registration/register.html')

        # Создаём пользователя
        user = User.objects.create_user(
            username=username,
            email=email,
            password=password
        )

        # Первый пользователь = создатель системы
        is_first_user = not User.objects.filter(is_staff=True, is_superuser=True).exclude(id=user.id).exists()

        if is_first_user:
            # Создатель системы — полный доступ
            user.is_staff = True
            user.is_superuser = True
            user.is_active = True
            admin_group, _ = Group.objects.get_or_create(name='Администратор')
            user.groups.add(admin_group)
            user.save()
            profile, _ = UserProfile.objects.get_or_create(user=user, defaults={'owner': None})
            if phone:
                profile.phone = phone
                profile.save()
        else:
            # Новая компания — ждёт одобрения создателем системы
            user.is_active = False
            user.is_staff = False
            user.is_superuser = False
            user.save()
            profile, _ = UserProfile.objects.get_or_create(user=user, defaults={'owner': None})
            if phone:
                profile.phone = phone
                profile.save()

        # Уведомление администратора в Telegram о новой регистрации
        try:
            if is_first_user:
                from apps.main.telegram_bot_complete import send_telegram_message
                admin_chat_id = getattr(settings, 'TELEGRAM_ADMIN_CHAT_ID', None)
                if admin_chat_id:
                    tg_text = (
                        f"🆕 <b>Новая регистрация</b>\n\n"
                        f"👤 Имя пользователя: <code>{username}</code>\n"
                        f"📧 Email: {email or '—'}\n"
                        f"🏷️ Роль: Администратор (первый пользователь)"
                    )
                    send_telegram_message(admin_chat_id, tg_text)
            else:
                import requests as req_lib
                bot_token = getattr(settings, 'TELEGRAM_BOT_TOKEN', None)
                chat_id = getattr(settings, 'TELEGRAM_ADMIN_CHAT_ID', None)
                if bot_token and chat_id:
                    phone_line = f"\n📞 Телефон: {phone}" if phone else ""
                    text = f"🏢 <b>Новая заявка на регистрацию компании</b>\n\n👤 Директор: <code>{user.username}</code>\n📧 Email: {user.email or '—'}{phone_line}\n\n⚠️ После одобрения — станет Директором новой компании."
                    keyboard = {
                        "inline_keyboard": [[
                            {"text": "✅ Одобрить", "callback_data": f"approve_{user.id}"},
                            {"text": "❌ Отклонить", "callback_data": f"reject_{user.id}"},
                        ]]
                    }
                    req_lib.post(
                        f"https://api.telegram.org/bot{bot_token}/sendMessage",
                        json={"chat_id": chat_id, "text": text, "parse_mode": "HTML", "reply_markup": keyboard},
                        timeout=5,
                    )
        except Exception:
            pass

        if is_first_user:
            # Первый пользователь — входим сразу
            login(request, user)
            messages.success(request, 'Добро пожаловать! Вы зарегистрированы как администратор.')
            return redirect('main:setup_company')
        else:
            # Обычный пользователь — ждёт одобрения
            return redirect('main:pending_approval')
    
    return render(request, 'registration/register.html')


def pending_approval(request):
    """Страница ожидания одобрения аккаунта"""
    return render(request, 'registration/pending_approval.html')


@login_required
def setup_company(request):
    """Страница настройки компании"""
    from apps.company.models import CompanyProfile
    
    company = CompanyProfile.get_company()
    
    if request.method == 'POST':
        # Обновляем данные компании
        company.company_name = request.POST.get('company_name')
        company.short_name = request.POST.get('short_name', '')
        company.phone = request.POST.get('phone', '')
        company.email = request.POST.get('email', '')
        company.address = request.POST.get('address', '')
        company.city = request.POST.get('city', '')
        company.inn = request.POST.get('inn', '')
        company.currency = request.POST.get('currency', 'сом')
        
        # Логотип
        if 'logo' in request.FILES:
            company.logo = request.FILES['logo']
        
        company.created_by = request.user
        company.save()
        
        messages.success(request, '✅ Профиль компании настроен!')
        return redirect('main:dashboard')
    
    context = {
        'company': company,
    }
    
    return render(request, 'main/setup_company.html', context)


@login_required
def edit_company(request):
    """Редактирование профиля компании"""
    from apps.company.models import CompanyProfile
    
    if not request.user.is_staff:
        messages.error(request, 'У вас нет прав на редактирование профиля компании')
        return redirect('main:dashboard')
    
    company = CompanyProfile.get_company()
    
    if request.method == 'POST':
        company.company_name = request.POST.get('company_name')
        company.short_name = request.POST.get('short_name', '')
        company.phone = request.POST.get('phone', '')
        company.email = request.POST.get('email', '')
        company.website = request.POST.get('website', '')
        company.address = request.POST.get('address', '')
        company.city = request.POST.get('city', '')
        company.inn = request.POST.get('inn', '')
        company.bank_account = request.POST.get('bank_account', '')
        company.bank_name = request.POST.get('bank_name', '')
        company.currency = request.POST.get('currency', 'сом')
        company.footer_text = request.POST.get('footer_text', '')
        
        if 'logo' in request.FILES:
            company.logo = request.FILES['logo']
        
        company.save()
        
        messages.success(request, '✅ Профиль компании обновлён!')
        return redirect('main:edit_company')
    
    context = {
        'company': company,
    }
    
    return render(request, 'main/edit_company.html', context)



def root_view(request):
    """Корневой URL: авторизованные → дашборд, неавторизованные → страница выбора роли"""
    if request.user.is_authenticated:
        return redirect('main:dashboard')
    return render(request, 'main/entry.html')


@login_required
def dashboard(request):
    """Современный дашборд с графиками (кэширование тяжёлых данных)"""
    from django.db.models import Sum
    from decimal import Decimal
    from django.core.cache import cache

    # Если новый директор ещё не настроил компанию — направить на setup
    if request.user.is_superuser and not request.user.is_staff:
        try:
            prof = request.user.profile
            if prof.needs_company_setup:
                prof.needs_company_setup = False
                prof.save()
                return redirect('main:setup_company')
        except Exception:
            pass

    owner = get_tenant_owner(request.user)
    now = timezone.now()

    # Принудительное обновление кэша по ?refresh=1
    force_refresh = request.GET.get('refresh') == '1'
    cache_key = f'dashboard_{owner.id}'
    CACHE_TTL = 120  # 2 минуты

    cached = None if force_refresh else cache.get(cache_key)
    if cached:
        ctx = cached
    else:
        week_ago = now - timedelta(days=7)
        month_ago = now - timedelta(days=30)

        # === СТАТИСТИКА ===
        total_clients = Client.objects.filter(owner=owner).count()
        new_clients_week = Client.objects.filter(owner=owner, created_at__gte=week_ago).count()
        active_orders = RentalOrder.objects.filter(status='open', client__owner=owner).count()
        total_orders = RentalOrder.objects.filter(client__owner=owner).count()

        # === ДОЛГИ ===
        all_clients = list(Client.objects.filter(owner=owner))
        total_debt = sum(float(c.get_debt()) for c in all_clients)
        debtors_count = sum(1 for c in all_clients if c.get_wallet_balance() < 0)
        total_credit = sum(float(c.get_credit()) for c in all_clients)

        # === ДОХОД ===
        total_payments = Payment.objects.filter(client__owner=owner).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        revenue_month = float(total_payments)
        total_paid = float(total_payments)

        prev_month_start = month_ago - timedelta(days=30)
        revenue_prev_month = Payment.objects.filter(
            client__owner=owner,
            payment_date__gte=prev_month_start,
            payment_date__lt=month_ago
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

        revenue_prev_float = float(revenue_prev_month)
        if revenue_prev_float > 0:
            revenue_growth = int(((revenue_month - revenue_prev_float) / revenue_prev_float) * 100)
        else:
            revenue_growth = 100 if revenue_month > 0 else 0

        # === ГРАФИК ДОХОДОВ ПО ДНЯМ ===
        revenue_labels = []
        revenue_data = []
        for i in range(6, -1, -1):
            day = now - timedelta(days=i)
            revenue_labels.append(day.strftime('%d.%m'))
            day_revenue = Payment.objects.filter(
                client__owner=owner,
                payment_date__date=day.date()
            ).aggregate(total=Sum('amount'))['total'] or 0
            revenue_data.append(float(day_revenue))

        # === ГРАФИК ЗАКАЗОВ ===
        orders_labels = []
        orders_open_data = []
        orders_closed_data = []
        for i in range(6, -1, -1):
            day = now - timedelta(days=i)
            orders_labels.append(day.strftime('%d.%m'))
            orders_open_data.append(RentalOrder.objects.filter(
                client__owner=owner, created_at__date=day.date(), status='open'
            ).count())
            orders_closed_data.append(RentalOrder.objects.filter(
                client__owner=owner, created_at__date=day.date(), status='closed'
            ).count())

        # === ТОП-5 ТОВАРОВ ===
        products_stats = OrderItem.objects.filter(order__client__owner=owner).values('product__name').annotate(
            rental_count=Count('id')
        ).order_by('-rental_count')[:5]
        max_count = products_stats[0]['rental_count'] if products_stats else 1
        top_products = [
            {
                'name': p['product__name'],
                'rental_count': p['rental_count'],
                'percentage': int((p['rental_count'] / max_count) * 100)
            }
            for p in products_stats
        ]

        # === ЗАКАЗЫ ПО ДНЯМ НЕДЕЛИ ===
        weekday_names = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']
        weekday_counts = [0] * 7
        for order in RentalOrder.objects.filter(client__owner=owner):
            weekday_counts[order.created_at.weekday()] += 1
        weekday_labels = json.dumps(weekday_names)
        weekday_data = json.dumps(weekday_counts)

        ctx = {
            'total_clients': total_clients,
            'new_clients_week': new_clients_week,
            'active_orders': active_orders,
            'total_orders': total_orders,
            'total_debt': total_debt,
            'debtors_count': debtors_count,
            'revenue_month': revenue_month,
            'revenue_growth': revenue_growth,
            'revenue_labels': json.dumps(revenue_labels),
            'revenue_data': json.dumps(revenue_data),
            'total_paid': total_paid,
            'total_credit': total_credit,
            'orders_labels': json.dumps(orders_labels),
            'orders_open_data': json.dumps(orders_open_data),
            'orders_closed_data': json.dumps(orders_closed_data),
            'top_products': top_products,
            'weekday_labels': weekday_labels,
            'weekday_data': weekday_data,
        }
        cache.set(cache_key, ctx, CACHE_TTL)

    # === ПРОСРОЧЕННЫЕ ЗАКАЗЫ (всегда свежие, не кэшируем) ===
    overdue_orders = []
    for order in RentalOrder.objects.filter(status='open', client__owner=owner).prefetch_related('items__product', 'client')[:50]:
        overdue_items_list = [
            item for item in order.items.all()
            if item.quantity_remaining > 0 and item.planned_return_date < now
        ]
        if overdue_items_list:
            days_overdue = (now - min(item.planned_return_date for item in overdue_items_list)).days
            overdue_orders.append({
                'id': order.id,
                'client': order.client,
                'days_overdue': days_overdue,
                'overdue_items': len(overdue_items_list),
                'total': order.get_current_total(),
            })
    overdue_orders.sort(key=lambda x: x['days_overdue'], reverse=True)

    ctx['now'] = now
    ctx['overdue_orders'] = overdue_orders

    return render(request, 'main/dashboard.html', ctx)

@login_required
def clients_list(request):
    """Список клиентов с фильтрами"""
    from apps.clients.models import Client

    owner = get_tenant_owner(request.user)

    # Параметры фильтров
    filter_type  = request.GET.get('filter', '')      # debt / credit / active / all
    date_range   = request.GET.get('date_range', '')  # today / week / month
    amount_min   = request.GET.get('amount_min', '')
    amount_max   = request.GET.get('amount_max', '')
    sort_by      = request.GET.get('sort', '-created_at')

    # Базовый queryset
    clients = Client.objects.filter(owner=owner).prefetch_related('phones', 'rental_orders')
    
    # === СОРТИРОВКА ===
    if sort_by == '-created_at':
        clients = clients.order_by('-created_at')
    elif sort_by == 'created_at':
        clients = clients.order_by('created_at')
    elif sort_by == 'name':
        clients = clients.order_by('last_name', 'first_name')
    else:
        clients = clients.order_by('-created_at')
    
    # === ФИЛЬТР ПО ДАТЕ РЕГИСТРАЦИИ ===
    now = timezone.now()
    if date_range == 'today':
        clients = clients.filter(created_at__date=now.date())
    elif date_range == 'week':
        clients = clients.filter(created_at__gte=now - timedelta(days=7))
    elif date_range == 'month':
        clients = clients.filter(created_at__gte=now - timedelta(days=30))
    
    # === ФИЛЬТР ПО ТИПУ (долг / аванс / активные) ===
    # Применяем post-filter (т.к. баланс вычисляется)
    if filter_type or amount_min or amount_max:
        filtered_ids = []
        for client in clients:
            balance = float(client.get_wallet_balance())
            debt = float(client.get_total_debt())
            
            # Фильтр по типу
            if filter_type == 'debt' and balance >= 0:
                continue
            if filter_type == 'credit' and balance <= 0:
                continue
            if filter_type == 'active':
                has_open = client.rental_orders.filter(status='open').exists()
                if not has_open:
                    continue
            
            # Фильтр по сумме долга
            if amount_min and debt < float(amount_min):
                continue
            if amount_max and debt > float(amount_max):
                continue
            
            filtered_ids.append(client.id)
        clients = clients.filter(id__in=filtered_ids)
    
    # Статистика
    all_clients = Client.objects.filter(owner=owner).prefetch_related('phones', 'rental_orders')
    debt_count   = sum(1 for c in all_clients if float(c.get_wallet_balance()) < 0)
    credit_count = sum(1 for c in all_clients if float(c.get_wallet_balance()) > 0)
    active_count = sum(1 for c in all_clients if c.rental_orders.filter(status='open').exists())
    
    # Пагинация
    page_obj, page_query = paginate(request, clients)

    context = {
        'clients': page_obj,
        'page_obj': page_obj,
        'page_query': page_query,
        'total_clients': all_clients.count(),
        'debt_count': debt_count,
        'credit_count': credit_count,
        'active_count': active_count,
        # Текущие фильтры
        'current_filter': filter_type,
        'current_date_range': date_range,
        'current_amount_min': amount_min,
        'current_amount_max': amount_max,
        'current_sort': sort_by,
        'filters_active': any([filter_type, date_range, amount_min, amount_max]),
    }

    return render(request, 'clients/list.html', context)

@login_required
def client_detail(request, client_id):
    """Карточка клиента с группированной историей"""
    import re as _re
    from django.utils import timezone

    owner = get_tenant_owner(request.user)
    client = get_object_or_404(Client, id=client_id, owner=owner)
    now = timezone.now()

    def _paid_for_order_cd(client_obj, order_id_int):
        """Сколько уже оплачено по конкретному заказу."""
        paid = 0.0
        for p in client_obj.payments.all():
            notes_text = p.notes or ''
            if 'Распределение:' in notes_text:
                for line in notes_text.split('\n'):
                    m = _re.search(
                        rf'Заказ\s*#{order_id_int}\s*:\s*(\d+(?:[\.,]\d+)?)\s*сом', line
                    )
                    if m:
                        paid += float(m.group(1).replace(',', '.'))
                        break
            elif (f'Заказ #{order_id_int}' in notes_text or f'#{order_id_int}' in notes_text):
                paid += float(p.amount)
        return paid

    active_orders = client.rental_orders.filter(status='open').prefetch_related('items__product')
    all_orders_qs = client.rental_orders.all().order_by('-created_at')
    payments = client.payments.all().order_by('-payment_date')[:10]

    # Добавляем per-order финансовые данные
    all_orders_data = []
    for order in all_orders_qs:
        total = float(order.get_current_total())
        base = float(order.get_original_total())
        overdue = total - base
        paid = _paid_for_order_cd(client, order.id)
        owed = max(0.0, total - paid)
        all_orders_data.append({
            'order': order,
            'total': total,
            'base': base,
            'overdue': overdue,
            'paid': paid,
            'owed': owed,
        })

    context = {
        'client': client,
        'balance': client.get_wallet_balance(),
        'debt': client.get_debt(),
        'credit': client.get_credit(),
        'total_paid': client.get_total_paid(),
        'total_debt': client.get_total_debt(),
        'active_orders': active_orders,
        'all_orders': all_orders_data,
        'payments': payments,
        'now': now,
    }

    return render(request, 'clients/detail.html', context)

@login_required
def client_payments(request, client_id):
    """Все платежи клиента"""
    client = get_object_or_404(Client, id=client_id)

    order_id = request.GET.get('order')

    # Получаем платежи клиента, при необходимости фильтруем по конкретному заказу
    payments_query = client.payments.all()
    if order_id:
        try:
            order_id_int = int(order_id)
        except (TypeError, ValueError):
            order_id_int = None
        if order_id_int:
            payments_query = payments_query.filter(notes__icontains=f'#{order_id_int}')
    payments = payments_query.order_by('-payment_date')

    context = {
        'client': client,
        'payments': payments,
        'total_paid': client.get_total_paid(),
        'wallet_balance': client.get_wallet_balance(),
        'order_id': order_id,
    }
    
    return render(request, 'payments/client_payments.html', context)

@manager_required
def create_order(request):
    """Создание нового заказа"""
    from apps.clients.models import Client
    from apps.inventory.models import Product
    from apps.rental.models import RentalOrder, OrderItem

    owner = get_tenant_owner(request.user)

    # GET запрос - показываем форму
    if request.method == 'GET':
        import json as _json
        from apps.inventory.models import Category
        clients = Client.objects.filter(owner=owner).prefetch_related('phones')
        categories = Category.objects.filter(owner=owner).order_by('name')
        # Build products JSON grouped by category for JS
        all_products = Product.objects.filter(owner=owner, is_active=True).select_related('category').order_by('name')
        products_by_cat = {}
        for p in all_products:
            cid = str(p.category_id)
            if cid not in products_by_cat:
                products_by_cat[cid] = []
            products_by_cat[cid].append({
                'id': p.id,
                'name': p.name,
                'available': p.quantity_available,
                'price_per_day': float(p.price_per_day),
                'price_per_hour': float(p.price_per_hour),
            })

        # Build clients JSON for search
        clients_json = []
        for c in clients:
            primary_phone = ''
            all_phones = []
            for ph in c.phones.all():
                all_phones.append(ph.phone_number)
                if ph.is_primary:
                    primary_phone = ph.phone_number
            clients_json.append({
                'id': c.id,
                'name': c.get_full_name(),
                'phone': primary_phone,
                'phones': all_phones,
            })

        preselect_client = request.GET.get('client', '')

        context = {
            'categories': categories,
            'clients_json': _json.dumps(clients_json),
            'products_by_cat_json': _json.dumps(products_by_cat),
            'preselect_client': preselect_client,
        }
        return render(request, 'rental/create_order.html', context)
    
    # POST запрос - обрабатываем форму
    if request.method == 'POST':
        try:
            import json as _json
            from decimal import Decimal as _Dec
            client_id = request.POST.get('client')
            # New format: cart_json = '[{"pid":1,"qty":2,"price":500,"rain":true},...]'
            cart_json = request.POST.get('cart_json', '[]')
            cart = _json.loads(cart_json)

            if not client_id:
                messages.error(request, 'Выберите клиента')
                return redirect('main:create_order')

            if not cart:
                messages.error(request, 'Добавьте хотя бы один товар')
                return redirect('main:create_order')

            client = Client.objects.get(id=client_id, owner=owner)

            # Global dates
            issued_date_str = request.POST.get('issued_date', '')
            planned_return_str = request.POST.get('planned_return_date', '')
            try:
                from django.utils.dateparse import parse_datetime
                issued_date = parse_datetime(issued_date_str) if issued_date_str else timezone.now()
                planned_return_date = parse_datetime(planned_return_str) if planned_return_str else None
                if issued_date and timezone.is_naive(issued_date):
                    issued_date = timezone.make_aware(issued_date)
                if planned_return_date and timezone.is_naive(planned_return_date):
                    planned_return_date = timezone.make_aware(planned_return_date)
            except Exception:
                issued_date = timezone.now()
                planned_return_date = None

            if not planned_return_date or planned_return_date <= issued_date:
                messages.error(request, 'Укажите корректные даты')
                return redirect('main:create_order')

            delta = planned_return_date - issued_date
            total_seconds = delta.total_seconds()
            global_days = int(total_seconds // 86400)
            global_hours = int((total_seconds % 86400) // 3600)

            proof_file = request.FILES.get('proof_file')
            has_delivery = request.POST.get('has_delivery') == 'on'
            delivery_cost_val = 0
            if has_delivery:
                try:
                    delivery_cost_val = float(request.POST.get('delivery_cost', '0') or '0')
                except (ValueError, TypeError):
                    delivery_cost_val = 0

            order = RentalOrder.objects.create(
                client=client,
                status='open',
                proof_file=proof_file,
                has_delivery=has_delivery,
                delivery_address=request.POST.get('delivery_address', '').strip() if has_delivery else '',
                delivery_vehicle=request.POST.get('delivery_vehicle', '').strip() if has_delivery else '',
                delivery_plate=request.POST.get('delivery_plate', '').strip() if has_delivery else '',
                delivery_cost=delivery_cost_val,
            )

            for item in cart:
                product_id = int(item['pid'])
                quantity = int(item['qty'])
                custom_price = float(item.get('price', 0))
                rain_applicable = bool(item.get('rain', False))
                by_hours = bool(item.get('byHours', False))

                product = Product.objects.get(id=product_id, owner=owner)

                if product.quantity_available < quantity:
                    messages.error(request, f'Недостаточно "{product.name}". Доступно: {product.quantity_available}')
                    order.delete()
                    return redirect('main:create_order')

                if by_hours:
                    # Price is per hour
                    price_per_hour = _Dec(str(custom_price)) if custom_price > 0 else (product.price_per_hour or product.price_per_day / 24)
                    price_per_day = product.price_per_day
                    total_hours = global_days * 24 + global_hours
                    original_cost = _Dec(str(quantity)) * price_per_hour * _Dec(str(total_hours))
                    # Store rental as 0 days + total_hours hours for clarity
                    item_rental_days = 0
                    item_rental_hours = total_hours
                else:
                    price_per_day = _Dec(str(custom_price)) if custom_price > 0 else product.price_per_day
                    price_per_hour = product.price_per_hour if product.price_per_hour else _Dec('0')
                    # Round up partial day
                    bill_days = global_days + (1 if global_hours > 0 else 0)
                    original_cost = _Dec(str(quantity)) * price_per_day * _Dec(str(bill_days))
                    item_rental_days = bill_days
                    item_rental_hours = 0

                OrderItem.objects.create(
                    order=order,
                    product=product,
                    quantity_taken=quantity,
                    quantity_remaining=quantity,
                    issued_date=issued_date,
                    planned_return_date=planned_return_date,
                    rental_days=item_rental_days,
                    rental_hours=item_rental_hours,
                    price_per_day=price_per_day,
                    price_per_hour=price_per_hour,
                    original_total_cost=original_cost,
                    current_total_cost=original_cost,
                    rain_applicable=rain_applicable,
                )
                product.update_available_quantity()
                product.quantity_available -= quantity
                product.save()
            
            messages.success(request, f'✅ Заказ #{order.id} успешно создан!')
            # Уведомляем директора
            try:
                from apps.main.telegram_bot_complete import notify_director_new_order
                notify_director_new_order(order)
            except Exception:
                pass
            # Email клиенту
            try:
                from apps.main.email_utils import notify_order_created_email
                notify_order_created_email(order)
            except Exception:
                pass
            log_activity(request.user, 'create_order', f'Создал заказ #{order.id} для клиента {client.get_full_name()}')
            from apps.main.cache_utils import invalidate_dashboard
            invalidate_dashboard(owner.id)
            if request.user != owner:
                from apps.main.notification_utils import push_to_owner
                push_to_owner(owner, f'📦 Новый заказ #{order.id}',
                              f'{request.user.username} создал заказ для {client.get_full_name()}',
                              type='order', link=f'/orders/{order.id}/')

            # Редирект на страницу клиента
            return redirect('main:client_detail', client_id=client.id)
            
        except Client.DoesNotExist:
            messages.error(request, 'Клиент не найден')
            return redirect('create_order')
        
        except Product.DoesNotExist:
            messages.error(request, 'Товар не найден')
            return redirect('create_order')
        
        except Exception as e:
            messages.error(request, f'Ошибка создания заказа: {e}')
            return redirect('create_order')
        
@login_required
def orders_list(request):
    """Список всех заказов с фильтрами"""
    from apps.rental.models import RentalOrder
    from apps.clients.models import Client

    owner = get_tenant_owner(request.user)

    # Получаем параметры фильтров
    status        = request.GET.get('status', '')        # open / closed / all
    date_range    = request.GET.get('date_range', '')    # today / week / month / all
    overdue_only  = request.GET.get('overdue', '')       # 1 = только просроченные
    amount_min    = request.GET.get('amount_min', '')    # минимальная сумма
    amount_max    = request.GET.get('amount_max', '')    # максимальная сумма
    sort_by       = request.GET.get('sort', '-created_at')  # сортировка

    # Базовый queryset — только заказы текущего тенанта
    orders = RentalOrder.objects.filter(client__owner=owner).select_related('client').prefetch_related('items')
    
    # === ФИЛЬТР ПО СТАТУСУ ===
    if status == 'open':
        orders = orders.filter(status='open')
    elif status == 'closed':
        orders = orders.filter(status='closed')
    
    # === ФИЛЬТР ПО ДАТЕ ===
    now = timezone.now()
    if date_range == 'today':
        orders = orders.filter(created_at__date=now.date())
    elif date_range == 'week':
        orders = orders.filter(created_at__gte=now - timedelta(days=7))
    elif date_range == 'month':
        orders = orders.filter(created_at__gte=now - timedelta(days=30))
    
    # === ФИЛЬТР ПРОСРОЧЕННЫХ ===
    if overdue_only == '1':
        overdue_ids = []
        for order in orders.filter(status='open'):
            for item in order.items.all():
                if item.quantity_remaining > 0 and item.planned_return_date < now:
                    overdue_ids.append(order.id)
                    break
        orders = orders.filter(id__in=overdue_ids)
    
    # === СОРТИРОВКА ===
    valid_sorts = ['-created_at', 'created_at', '-id', 'id']
    if sort_by in valid_sorts:
        orders = orders.order_by(sort_by)
    else:
        orders = orders.order_by('-created_at')
    
    # === ФИЛЬТР ПО СУММЕ (post-filter т.к. сумма вычисляется) ===
    if amount_min or amount_max:
        filtered = []
        for order in orders:
            total = float(order.get_current_total())
            if amount_min and total < float(amount_min):
                continue
            if amount_max and total > float(amount_max):
                continue
            filtered.append(order.id)
        orders = orders.filter(id__in=filtered)
    
    # Статистика
    total_open   = RentalOrder.objects.filter(status='open', client__owner=owner).count()
    total_closed = RentalOrder.objects.filter(status='closed', client__owner=owner).count()

    # Просроченные
    overdue_count = 0
    for order in RentalOrder.objects.filter(status='open', client__owner=owner).prefetch_related('items'):
        for item in order.items.all():
            if item.quantity_remaining > 0 and item.planned_return_date < now:
                overdue_count += 1
                break
    
    # Пагинация
    page_obj, page_query = paginate(request, orders)

    context = {
        'orders': page_obj,
        'page_obj': page_obj,
        'page_query': page_query,
        'total_open': total_open,
        'total_closed': total_closed,
        'overdue_count': overdue_count,
        # Текущие фильтры (для отображения в форме)
        'current_status': status,
        'current_date_range': date_range,
        'current_overdue': overdue_only,
        'current_amount_min': amount_min,
        'current_amount_max': amount_max,
        'current_sort': sort_by,
        # Флаг что фильтр активен
        'filters_active': any([status, date_range, overdue_only, amount_min, amount_max]),
    }

    return render(request, 'rental/orders_list.html', context)


@cashier_required

def accept_payment(request):
    """Принять оплату с отслеживанием заказов"""

    import re

    def _distribution_amount_for_order(notes_text: str, order_id_int: int):
        if not notes_text:
            return None
        if 'Распределение:' in notes_text:
            for line in notes_text.split('\n'):
                pattern = rf'Заказ\s*#{order_id_int}\s*:\s*(\d+(?:[\.,]\d+)?)\s*сом'
                match = re.search(pattern, line)
                if match:
                    return float(match.group(1).replace(',', '.'))
        return None

    def _paid_for_order(client_obj, order_id_int: int) -> float:
        paid = 0.0
        for p in client_obj.payments.all():
            amt = _distribution_amount_for_order(p.notes or '', order_id_int)
            if amt is not None:
                paid += amt
                continue

            # Фолбэк для старых платежей без блока "Распределение"
            if p.notes and (f'Заказ #{order_id_int}' in p.notes or f'#{order_id_int}' in p.notes):
                # Если явно указан заказ в тексте — считаем весь платёж относящимся к нему
                paid += float(p.amount)
        return paid
    
    if request.method == 'POST':
        client_id = request.POST.get('client_id')
        amount = request.POST.get('amount')
        payment_method = request.POST.get('payment_method', 'cash')
        notes = request.POST.get('notes', '')
        selected_orders = request.POST.getlist('selected_orders')
        
        try:
            client = get_object_or_404(Client, id=client_id)
            payment_amount = float(amount)
            
            if payment_amount <= 0:
                return redirect('main:accept_payment')
            
            # Определяем заказы для оплаты
            if selected_orders:
                orders_to_pay = RentalOrder.objects.filter(
                    id__in=selected_orders,
                    client=client,
                ).order_by('created_at')
            else:
                # Важно: распределяем не только по open, но и по closed, если там остался долг
                orders_to_pay = client.rental_orders.all().order_by('created_at')
            
            # Распределяем оплату
            remaining_payment = payment_amount
            payment_distribution = []
            paid_orders = []
            
            for order in orders_to_pay:
                if remaining_payment <= 0:
                    break

                already_paid = _paid_for_order(client, order.id)
                order_due = float(order.get_current_total()) - already_paid
                if order_due <= 0:
                    continue

                if remaining_payment >= order_due:
                    # Полностью оплачен
                    paid_amount = order_due
                    remaining_payment -= order_due
                    payment_distribution.append(f'Заказ #{order.id}: {paid_amount:.0f} сом (полностью)')
                    paid_orders.append(str(order.id))
                else:
                    # Частично оплачен
                    paid_amount = remaining_payment
                    payment_distribution.append(f'Заказ #{order.id}: {paid_amount:.0f} сом (частично)')
                    paid_orders.append(str(order.id))
                    remaining_payment = 0
            
            # Формируем примечание
            if paid_orders:
                order_note = f"Оплата для заказ{'ов' if len(paid_orders) > 1 else 'а'} #{', #'.join(paid_orders)}"
                if notes:
                    full_notes = f"{order_note}\n{notes}\n\nРаспределение:\n" + '\n'.join(payment_distribution)
                else:
                    full_notes = f"{order_note}\n\nРаспределение:\n" + '\n'.join(payment_distribution)

                # Если остался остаток — это аванс
                if remaining_payment > 0:
                    full_notes += f"\n\nАванс: {remaining_payment:.0f} сом"
            else:
                full_notes = notes if notes else 'Аванс (нет долгов по заказам)'
            
            # Создаём оплату
            payment_obj = Payment.objects.create(
                client=client,
                amount=payment_amount,
                payment_method=payment_method,
                notes=full_notes
            )
            log_activity(request.user, 'accept_payment', f'Принял оплату {payment_amount:.0f} сом от клиента {client.get_full_name()}')
            # Уведомляем директора
            try:
                from apps.main.telegram_bot_complete import notify_director_payment
                notify_director_payment(payment_obj)
            except Exception:
                pass
            _owner = get_tenant_owner(request.user)
            from apps.main.cache_utils import invalidate_dashboard
            invalidate_dashboard(_owner.id)
            if request.user != _owner:
                from apps.main.notification_utils import push_to_owner
                push_to_owner(_owner, f'💰 Оплата {payment_amount:.0f} сом',
                              f'{request.user.username} принял оплату от {client.get_full_name()}',
                              type='payment', link='/payment/')

            return redirect('main:client_detail', client_id=client.id)

        except (ValueError, TypeError):
            return redirect('main:accept_payment')
    
    # GET запрос
    owner = get_tenant_owner(request.user)
    clients = Client.objects.filter(owner=owner).order_by('last_name', 'first_name')
    
    # Клиенты с долгом
    clients_with_debt = []
    for client in clients:
        debt = client.get_debt()
        if debt > 0:
            open_orders = client.rental_orders.all().order_by('created_at')
            orders_list = []
            for order in open_orders:
                already_paid = _paid_for_order(client, order.id)
                order_due = float(order.get_current_total()) - already_paid
                if order_due <= 0:
                    continue
                orders_list.append({
                    'id': order.id,
                    'created_at': order.created_at,
                    'total': float(order.get_current_total()),
                    'due': order_due,
                    'status': order.status,
                    'items_count': order.items.count(),
                })
            
            clients_with_debt.append({
                'client': client,
                'debt': debt,
                'orders': orders_list,
            })
    
    # Выбранный клиент
    selected_client_id = request.GET.get('client')
    selected_client = None
    selected_client_orders = []
    
    if selected_client_id:
        try:
            selected_client = Client.objects.get(id=selected_client_id)
            for order in selected_client.rental_orders.all().order_by('created_at'):
                already_paid = _paid_for_order(selected_client, order.id)
                order_due = float(order.get_current_total()) - already_paid
                if order_due <= 0:
                    continue
                selected_client_orders.append({
                    'id': order.id,
                    'created_at': order.created_at,
                    'total': float(order.get_current_total()),
                    'due': order_due,
                    'status': order.status,
                    'items_count': order.items.count(),
                })
        except Client.DoesNotExist:
            pass
    
    clients_with_debt_ids = [d['client'].id for d in clients_with_debt]

    context = {
        'clients': clients,
        'clients_with_debt': clients_with_debt,
        'clients_with_debt_ids': clients_with_debt_ids,
        'selected_client': selected_client,
        'selected_client_id': selected_client_id,
        'selected_client_orders': selected_client_orders,
    }
    
    return render(request, 'rental/payment.html', context)


@login_required
def client_orders_json(request, client_id):
    """AJAX: вернуть список заказов клиента с суммами к оплате"""
    import json as _json
    client = get_object_or_404(Client, id=client_id)

    def _paid_for_order_local(client_obj, order_id_int):
        import re
        paid = 0.0
        for p in client_obj.payments.all():
            notes_text = p.notes or ''
            if 'Распределение:' in notes_text:
                for line in notes_text.split('\n'):
                    m = re.search(rf'Заказ\s*#{order_id_int}\s*:\s*(\d+(?:[\.,]\d+)?)\s*сом', line)
                    if m:
                        paid += float(m.group(1).replace(',', '.'))
                        break
            elif f'Заказ #{order_id_int}' in notes_text or f'#{order_id_int}' in notes_text:
                paid += float(p.amount)
        return paid

    orders_data = []
    for order in client.rental_orders.filter(status='open').order_by('created_at'):
        already_paid = _paid_for_order_local(client, order.id)
        due = float(order.get_current_total()) - already_paid
        if due <= 0:
            continue
        orders_data.append({
            'id': order.id,
            'created_at': order.created_at.strftime('%d.%m.%Y %H:%M'),
            'items_count': order.items.count(),
            'status': order.status,
            'status_display': 'открыт' if order.status == 'open' else 'закрыт',
            'due': round(due),
            'total': round(float(order.get_current_total())),
        })

    return JsonResponse({'orders': orders_data})


@login_required
def returns_page(request):
    """Страница возвратов с предложением оплаты"""
    
    if request.method == 'POST':
        from apps.rental.models import ReturnDocument, ReturnItem
        from decimal import Decimal

        client_id = request.POST.get('client_id')
        notes = request.POST.get('notes', '')
        payment_amount = request.POST.get('payment_amount', '0')
        payment_method = request.POST.get('payment_method', 'cash')
        payment_order_id = request.POST.get('payment_order_id', '')  # ручной выбор заказа

        # Создаём документ возврата
        return_doc = ReturnDocument.objects.create(notes=notes)

        # Обрабатываем возвраты
        has_returns = False
        total_cost = 0
        total_repair = 0
        returned_order_ids = []  # список order_id в порядке обработки

        for key in request.POST:
            if key.startswith('return_') and request.POST[key]:
                item_id = key.replace('return_', '')
                try:
                    quantity = int(request.POST[key])
                except (ValueError, TypeError):
                    continue

                if quantity > 0:
                    try:
                        order_item = OrderItem.objects.get(id=item_id)

                        if quantity <= order_item.quantity_remaining:
                            repair_fee_val = request.POST.get(f'repair_fee_{item_id}', '0') or '0'
                            repair_notes_val = request.POST.get(f'repair_notes_{item_id}', '')
                            try:
                                repair_fee_dec = Decimal(str(float(repair_fee_val)))
                            except (ValueError, TypeError):
                                repair_fee_dec = Decimal('0')

                            return_item = ReturnItem(
                                return_document=return_doc,
                                order_item=order_item,
                                quantity=quantity,
                                repair_fee=repair_fee_dec,
                                repair_notes=repair_notes_val,
                            )
                            # calculated_cost, actual_days/hours set in save()
                            return_item.save()
                            has_returns = True
                            total_cost += float(return_item.calculated_cost)
                            total_repair += float(repair_fee_dec)

                            # Запоминаем заказ
                            oid = order_item.order_id
                            if oid not in returned_order_ids:
                                returned_order_ids.append(oid)

                            # ✅ ВАЖНО: ОБНОВЛЯЕМ ДОСТУПНОСТЬ ТОВАРА
                            product = order_item.product
                            product.update_available_quantity()

                    except OrderItem.DoesNotExist:
                        pass

        if has_returns:
            # Если клиент оплатил — распределяем платёж по заказам
            try:
                payment_amount_float = float(payment_amount)
            except (ValueError, TypeError):
                payment_amount_float = 0

            if payment_amount_float > 0:
                import re as _re

                def _dist_amt(notes_text, oid):
                    if notes_text and 'Распределение:' in notes_text:
                        for ln in notes_text.split('\n'):
                            m = _re.search(rf'Заказ\s*#{oid}\s*:\s*(\d+(?:[\.,]\d+)?)\s*сом', ln)
                            if m:
                                return float(m.group(1).replace(',', '.'))
                    return None

                def _paid_for(client_obj, oid):
                    total = 0.0
                    for p in client_obj.payments.all():
                        amt = _dist_amt(p.notes or '', oid)
                        if amt is not None:
                            total += amt
                        elif p.notes and (f'Заказ #{oid}' in p.notes or f'#{oid}' in p.notes):
                            total += float(p.amount)
                    return total

                client_obj = Client.objects.get(id=client_id)

                # Если указан конкретный заказ — только он, иначе заказы с возвратами
                if payment_order_id:
                    try:
                        manual_order = RentalOrder.objects.get(id=payment_order_id, client=client_obj)
                        orders_for_payment = [manual_order]
                    except RentalOrder.DoesNotExist:
                        orders_for_payment = list(RentalOrder.objects.filter(
                            id__in=returned_order_ids, client=client_obj
                        ).order_by('created_at'))
                else:
                    orders_for_payment = list(RentalOrder.objects.filter(
                        id__in=returned_order_ids, client=client_obj
                    ).order_by('created_at'))

                remaining = payment_amount_float
                distribution = []
                paid_ids = []

                for order in orders_for_payment:
                    if remaining <= 0:
                        break
                    already = _paid_for(client_obj, order.id)
                    due = float(order.get_current_total()) - already
                    if due <= 0:
                        continue
                    if remaining >= due:
                        distribution.append(f'Заказ #{order.id}: {due:.0f} сом (полностью)')
                        paid_ids.append(str(order.id))
                        remaining -= due
                    else:
                        distribution.append(f'Заказ #{order.id}: {remaining:.0f} сом (частично)')
                        paid_ids.append(str(order.id))
                        remaining = 0

                if paid_ids:
                    order_word = 'ов' if len(paid_ids) > 1 else 'а'
                    full_notes = (
                        f"Оплата при возврате (Возврат #{return_doc.id}) "
                        f"для заказ{order_word} #{', #'.join(paid_ids)}"
                        f"\n\nРаспределение:\n" + '\n'.join(distribution)
                    )
                    if remaining > 0:
                        full_notes += f"\n\nАванс: {remaining:.0f} сом"
                else:
                    full_notes = f'Оплата при возврате (Возврат #{return_doc.id})'

                Payment.objects.create(
                    client_id=client_id,
                    amount=payment_amount_float,
                    payment_method=payment_method,
                    notes=full_notes
                )
            
            if total_repair > 0:
                messages.success(request, f'✅ Товары возвращены! Плата за ремонт/чистку: {int(total_repair)} сом')
            else:
                messages.success(request, '✅ Товары возвращены и инвентарь обновлён!')
            return redirect('main:client_detail', client_id=client_id)
        else:
            return_doc.delete()
            messages.warning(request, '⚠️ Нет товаров для возврата')
            return redirect(f'/rental/returns/?client_id={client_id}')
    
    # GET запрос
    _returns_owner = get_tenant_owner(request.user)
    clients_with_orders = []

    for client in Client.objects.filter(owner=_returns_owner).prefetch_related('phones'):
        active_orders = client.rental_orders.filter(status='open')
        if active_orders.exists():
            has_items = False
            for order in active_orders:
                if order.items.filter(quantity_remaining__gt=0).exists():
                    has_items = True
                    break

            if has_items:
                phones = [p.phone_number for p in client.phones.all()]
                clients_with_orders.append({
                    'id': client.id,
                    'get_full_name': client.get_full_name(),
                    'active_count': active_orders.count(),
                    'balance': int(client.get_wallet_balance()),
                    'phones': phones,
                    'phones_str': ' '.join(phones),
                })
    
    # Если выбран клиент
    client_id = request.GET.get('client_id')
    selected_orders = []
    selected_client = None
    
    if client_id:
        try:
            import re as _re2
            from django.utils import timezone as _tz
            _now = _tz.now()
            selected_client = Client.objects.get(id=client_id)

            def _paid_for_order_ret(client_obj, order_id_int):
                paid = 0.0
                for p in client_obj.payments.all():
                    notes = p.notes or ''
                    if 'Распределение:' in notes:
                        for ln in notes.split('\n'):
                            m = _re2.search(rf'Заказ\s*#{order_id_int}\s*:\s*(\d+(?:[\.,]\d+)?)\s*сом', ln)
                            if m:
                                paid += float(m.group(1).replace(',', '.'))
                                break
                    elif f'Заказ #{order_id_int}' in notes or f'#{order_id_int}' in notes:
                        paid += float(p.amount)
                return paid

            for order in selected_client.rental_orders.filter(status='open'):
                items_data = []
                for item in order.items.filter(quantity_remaining__gt=0):
                    _overdue = item.quantity_remaining > 0 and item.planned_return_date < _now
                    _overdue_delta = (_now - item.planned_return_date) if _overdue else None
                    items_data.append({
                        'id': item.id,
                        'product_name': item.product.name,
                        'quantity_taken': item.quantity_taken,
                        'quantity_returned': item.quantity_returned,
                        'quantity_remaining': item.quantity_remaining,
                        'issued_date': item.issued_date.strftime('%d.%m.%Y %H:%M'),
                        'planned_return_date': item.planned_return_date.strftime('%d.%m.%Y %H:%M'),
                        'current_total_cost': float(item.current_total_cost),
                        'actual_cost': float(item.get_actual_cost()),
                        'price_per_day': float(item.price_per_day),
                        'rental_days': item.rental_days,
                        'is_overdue': _overdue,
                        'overdue_days': _overdue_delta.days if _overdue_delta else 0,
                        'overdue_hours': (_overdue_delta.seconds // 3600) if _overdue_delta else 0,
                        'remaining_planned': round(float(item.price_per_day) * item.rental_days * item.quantity_remaining, 0),
                        'overdue_extra': round(float(item.price_per_day) * (_overdue_delta.days if _overdue_delta else 0) * item.quantity_remaining, 0),
                    })

                if items_data:
                    _total = float(order.get_current_total()) + float(order.delivery_cost)
                    _paid = _paid_for_order_ret(selected_client, order.id)
                    _remaining = max(0.0, _total - _paid)
                    selected_orders.append({
                        'id': order.id,
                        'order_code': order.order_code,
                        'created_at': order.created_at.strftime('%d.%m.%Y %H:%M'),
                        'total': _total,
                        'paid': _paid,
                        'remaining': _remaining,
                        'delivery_cost': float(order.delivery_cost),
                        'items': items_data,
                    })
        except Client.DoesNotExist:
            pass
    
    context = {
        'clients_with_orders': clients_with_orders,
        'selected_orders': selected_orders,
        'selected_client_id': client_id,
        'selected_client': selected_client,
    }
    
    return render(request, 'rental/returns.html', context)

@login_required
def history(request):
    """История операций сгруппированная по заказам"""
    from django.utils import timezone

    owner = get_tenant_owner(request.user)
    selected_client_id = request.GET.get('client_id', '')
    clients = Client.objects.filter(owner=owner).order_by('last_name', 'first_name')

    if not selected_client_id:
        context = {
            'clients': clients,
            'selected_client': None,
            'order_groups': [],
        }
        return render(request, 'main/history.html', context)

    try:
        selected_client = Client.objects.get(id=selected_client_id, owner=owner)
    except Client.DoesNotExist:
        return redirect('main:history')
    
    now = timezone.now()
    order_groups = get_order_groups_for_client(selected_client, now)
    
    context = {
        'clients': clients,
        'selected_client': selected_client,
        'selected_client_id': selected_client_id,
        'order_groups': order_groups,
        'wallet_balance': selected_client.get_wallet_balance(),
        'total_paid': selected_client.get_total_paid(),
        'total_debt': selected_client.get_total_debt(),
    }
    
    return render(request, 'main/history.html', context)
@manager_required
def edit_order(request, order_id):
    """Редактирование заказа"""
    
    order = get_object_or_404(RentalOrder, id=order_id)
    
    # Только открытые заказы можно редактировать
    if order.status != 'open':
        return redirect('main:client_detail', client_id=order.client.id)
    
    if request.method == 'POST':
        from decimal import Decimal
        from datetime import timedelta
        import re as _re
        _log_pattern = _re.compile(r'^\[\d{2}\.\d{2}\.\d{4} \d{2}:\d{2}\]')
        _local_now = timezone.localtime(timezone.now())
        _username = request.user.username

        # === ДОБАВЛЕНИЕ НОВОГО ТОВАРА ===
        if 'add_product' in request.POST:
            product_id = request.POST.get('new_product_id')
            quantity = int(request.POST.get('new_quantity', 0))
            days = int(request.POST.get('new_days', 0))
            hours = int(request.POST.get('new_hours', 0))
            
            if product_id and quantity > 0 and (days > 0 or hours > 0):
                try:
                    product = Product.objects.get(id=product_id)
                    
                    # Проверяем доступность
                    if product.quantity_available >= quantity:
                        # Уменьшаем доступное количество
                        product.quantity_available -= quantity
                        product.save()
                        
                        # Создаём новый OrderItem
                        issued_date = order.created_at
                        planned_return_date = issued_date + timedelta(days=days, hours=hours)
                        
                        # Рассчитываем стоимость
                        total_rental_hours = (days * 24) + hours
                        total_days_decimal = Decimal(total_rental_hours) / Decimal(24)
                        total_cost = product.price_per_day * total_days_decimal * Decimal(quantity)
                        
                        OrderItem.objects.create(
                            order=order,
                            product=product,
                            quantity_taken=quantity,
                            quantity_returned=0,
                            quantity_remaining=quantity,
                            rental_days=days,
                            rental_hours=hours,
                            price_per_day=product.price_per_day,
                            price_per_hour=product.price_per_hour,
                            issued_date=issued_date,
                            planned_return_date=planned_return_date,
                            original_total_cost=total_cost,
                            current_total_cost=total_cost,
                        )
                        product.update_available_quantity()
                        # Log добавление товара
                        _log_line = (
                            f"[{_local_now.strftime('%d.%m.%Y %H:%M')}] "
                            f"{_username}: "
                            f"Добавил товар {product.name} x{quantity} на {days} дн {hours} ч ({int(total_cost)} сом)"
                        )
                        if order.notes:
                            order.notes += '\n' + _log_line
                        else:
                            order.notes = _log_line
                        order.save()
                        return redirect('main:edit_order', order_id=order.id)
                except Product.DoesNotExist:
                    pass
        
        # === ОБНОВЛЕНИЕ ПРИМЕЧАНИЙ ===
        # Сохраняем пользовательские заметки, сохраняя системные лог-строки
        _user_notes = request.POST.get('notes', '').strip()
        _system_log_lines = []
        if order.notes:
            for _line in order.notes.split('\n'):
                if _log_pattern.match(_line.strip()):
                    _system_log_lines.append(_line)
        _combined = _user_notes
        if _system_log_lines:
            if _combined:
                _combined += '\n\n'
            _combined += '\n'.join(_system_log_lines)
        order.notes = _combined
        order.save()

        # Список новых лог-записей для этого редактирования
        _edit_logs = []

        # === ОБНОВЛЕНИЕ ТОВАРОВ ===
        for key in request.POST:
            if key.startswith('item_quantity_'):
                item_id = key.replace('item_quantity_', '')
                new_quantity = int(request.POST.get(key, 0))
                
                try:
                    order_item = OrderItem.objects.get(id=item_id, order=order)
                    
                    if new_quantity >= order_item.quantity_returned:
                        old_quantity = order_item.quantity_taken
                        difference = new_quantity - old_quantity
                        
                        if difference > 0:
                            if order_item.product.quantity_available >= difference:
                                order_item.product.quantity_available -= difference
                                order_item.product.save()

                                order_item.quantity_taken = new_quantity
                                order_item.quantity_remaining = new_quantity - order_item.quantity_returned

                                # Пересчёт
                                total_rental_hours = (order_item.rental_days * 24) + order_item.rental_hours
                                total_days_decimal = Decimal(total_rental_hours) / Decimal(24)
                                order_item.original_total_cost = order_item.price_per_day * total_days_decimal * Decimal(new_quantity)
                                order_item.current_total_cost = order_item.original_total_cost

                                order_item.save()
                                _edit_logs.append(
                                    f"[{_local_now.strftime('%d.%m.%Y %H:%M')}] {_username}: "
                                    f"Изменил количество {order_item.product.name} с {old_quantity} на {new_quantity} шт"
                                )
                        else:
                            if difference != 0:
                                order_item.product.quantity_available += abs(difference)
                                order_item.product.save()

                            order_item.quantity_taken = new_quantity
                            order_item.quantity_remaining = new_quantity - order_item.quantity_returned

                            # Пересчёт
                            total_rental_hours = (order_item.rental_days * 24) + order_item.rental_hours
                            total_days_decimal = Decimal(total_rental_hours) / Decimal(24)
                            order_item.original_total_cost = order_item.price_per_day * total_days_decimal * Decimal(new_quantity)
                            order_item.current_total_cost = order_item.original_total_cost

                            order_item.save()
                            if difference != 0:
                                _edit_logs.append(
                                    f"[{_local_now.strftime('%d.%m.%Y %H:%M')}] {_username}: "
                                    f"Изменил количество {order_item.product.name} с {old_quantity} на {new_quantity} шт"
                                )
                
                except OrderItem.DoesNotExist:
                    pass
            
            # Обновление дней
            if key.startswith('item_days_'):
                item_id = key.replace('item_days_', '')
                new_days = int(request.POST.get(key, 0))
                
                try:
                    order_item = OrderItem.objects.get(id=item_id, order=order)
                    old_days = order_item.rental_days
                    order_item.rental_days = new_days
                    order_item.planned_return_date = order_item.issued_date + timedelta(days=new_days, hours=order_item.rental_hours)

                    # Пересчёт
                    total_rental_hours = (new_days * 24) + order_item.rental_hours
                    total_days_decimal = Decimal(total_rental_hours) / Decimal(24)
                    order_item.original_total_cost = order_item.price_per_day * total_days_decimal * Decimal(order_item.quantity_taken)
                    order_item.current_total_cost = order_item.original_total_cost

                    order_item.save()
                    if old_days != new_days:
                        _edit_logs.append(
                            f"[{_local_now.strftime('%d.%m.%Y %H:%M')}] {_username}: "
                            f"Изменил срок {order_item.product.name} с {old_days} дн на {new_days} дн"
                        )
                except OrderItem.DoesNotExist:
                    pass

            # Обновление часов
            if key.startswith('item_hours_'):
                item_id = key.replace('item_hours_', '')
                new_hours = int(request.POST.get(key, 0))
                
                try:
                    order_item = OrderItem.objects.get(id=item_id, order=order)
                    old_hours = order_item.rental_hours
                    order_item.rental_hours = new_hours
                    order_item.planned_return_date = order_item.issued_date + timedelta(days=order_item.rental_days, hours=new_hours)

                    # Пересчёт
                    total_rental_hours = (order_item.rental_days * 24) + new_hours
                    total_days_decimal = Decimal(total_rental_hours) / Decimal(24)
                    order_item.original_total_cost = order_item.price_per_day * total_days_decimal * Decimal(order_item.quantity_taken)
                    order_item.current_total_cost = order_item.original_total_cost

                    order_item.save()
                    if old_hours != new_hours:
                        _edit_logs.append(
                            f"[{_local_now.strftime('%d.%m.%Y %H:%M')}] {_username}: "
                            f"Изменил часы {order_item.product.name} с {old_hours} ч на {new_hours} ч"
                        )
                except OrderItem.DoesNotExist:
                    pass
        
        # === УДАЛЕНИЕ ТОВАРОВ ===
        for key in request.POST:
            if key.startswith('delete_item_'):
                item_id = key.replace('delete_item_', '')

                try:
                    order_item = OrderItem.objects.get(id=item_id, order=order)

                    if order_item.quantity_returned == 0:
                        _deleted_name = order_item.product.name
                        _deleted_qty = order_item.quantity_taken
                        order_item.product.quantity_available += order_item.quantity_taken
                        order_item.product.save()
                        order_item.delete()
                        _edit_logs.append(
                            f"[{_local_now.strftime('%d.%m.%Y %H:%M')}] {_username}: "
                            f"Удалил товар {_deleted_name} x{_deleted_qty} шт"
                        )

                except OrderItem.DoesNotExist:
                    pass

        # Сохраняем новые лог-записи в notes заказа
        if _edit_logs:
            _log_block = '\n'.join(_edit_logs)
            order.refresh_from_db()
            if order.notes:
                order.notes += '\n' + _log_block
            else:
                order.notes = _log_block
            order.save()

        return redirect('main:client_detail', client_id=order.client.id)
    
    # GET запрос
    import re as _re
    _log_pat = _re.compile(r'^\[\d{2}\.\d{2}\.\d{4} \d{2}:\d{2}\]')
    user_notes_display = ''
    if order.notes:
        non_system = [l for l in order.notes.split('\n') if not _log_pat.match(l.strip())]
        user_notes_display = '\n'.join(non_system).strip()

    _edit_owner = get_tenant_owner(request.user)
    available_products = Product.objects.filter(owner=_edit_owner, is_active=True, quantity_available__gt=0).order_by('name')

    # --- Данные для дождливого календаря ---
    from apps.rental.models import OrderExcludedDay
    from datetime import date as date_cls, timedelta as td
    import json as _json
    items_all = list(order.items.all())
    cal_start = cal_end = None
    for _it in items_all:
        if _it.issued_date:
            _d = _it.issued_date.date() if hasattr(_it.issued_date, 'date') else _it.issued_date
            cal_start = _d if cal_start is None else min(cal_start, _d)
        if _it.planned_return_date:
            _d = _it.planned_return_date.date() if hasattr(_it.planned_return_date, 'date') else _it.planned_return_date
            cal_end = _d if cal_end is None else max(cal_end, _d)
    # Все даты в диапазоне (максимум 90 дней)
    calendar_days = []
    if cal_start and cal_end:
        cur = cal_start
        while cur <= cal_end and (cur - cal_start).days < 91:
            calendar_days.append(cur)
            cur += td(days=1)
    excluded_dates = [str(d) for d in order.excluded_days.values_list('date', flat=True)]
    rain_total = order.get_total_excluding_rain()
    # ---

    context = {
        'order': order,
        'available_products': available_products,
        'user_notes_display': user_notes_display,
        'calendar_days': calendar_days,
        'excluded_dates': excluded_dates,
        'excluded_dates_json': _json.dumps(excluded_dates),
        'rain_excluded_count': len(excluded_dates),
        'rain_total': rain_total,
    }

    return render(request, 'rental/edit_order.html', context)

@login_required
def reset_client_balance(request, client_id):
    """Списать долг клиента (обнулить баланс). Только для суперпользователя."""
    import re as _re

    if not request.user.is_superuser:
        messages.error(request, 'Нет прав для этого действия.')
        return redirect('main:client_detail', client_id=client_id)

    if request.method != 'POST':
        return redirect('main:client_detail', client_id=client_id)

    client = get_object_or_404(Client, id=client_id)
    balance = round(client.get_wallet_balance(), 2)

    if abs(balance) < 0.005:
        messages.info(request, 'Баланс клиента уже нулевой.')
        return redirect('main:client_detail', client_id=client_id)

    if balance > 0:
        # Аванс — записываем отрицательный платёж для обнуления
        Payment.objects.create(
            client=client,
            amount=-balance,
            payment_method='writeoff',
            notes=f'Списание аванса (обнуление баланса): -{balance:.0f} сом',
        )
        log_activity(request.user, 'reset_balance',
                     f'Обнулил аванс клиента {client.get_full_name()} (списано {balance:.0f} сом)')
        messages.success(request, f'Аванс {balance:.0f} сом списан. Баланс обнулён.')
        return redirect('main:client_detail', client_id=client_id)

    # Долг — записываем положительный платёж-списание с распределением по заказам
    debt = abs(balance)

    def _dist_amt(notes_text, oid):
        if notes_text and 'Распределение:' in notes_text:
            for ln in notes_text.split('\n'):
                m = _re.search(rf'Заказ\s*#{oid}\s*:\s*(\d+(?:[\.,]\d+)?)\s*сом', ln)
                if m:
                    return float(m.group(1).replace(',', '.'))
        return None

    def _paid_for(c, oid):
        total = 0.0
        for p in c.payments.all():
            amt = _dist_amt(p.notes or '', oid)
            if amt is not None:
                total += amt
            elif p.notes and (f'Заказ #{oid}' in p.notes or f'#{oid}' in p.notes):
                total += float(p.amount)
        return total

    orders = list(client.rental_orders.all().order_by('created_at'))
    remaining = debt
    distribution = []
    paid_ids = []

    for order in orders:
        if remaining <= 0:
            break
        already = _paid_for(client, order.id)
        due = round(float(order.get_current_total()) - already, 2)
        if due <= 0:
            continue
        if remaining >= due:
            distribution.append(f'Заказ #{order.id}: {due:.0f} сом (полностью)')
            paid_ids.append(str(order.id))
            remaining -= due
        else:
            distribution.append(f'Заказ #{order.id}: {remaining:.0f} сом (частично)')
            paid_ids.append(str(order.id))
            remaining = 0

    if paid_ids:
        order_word = 'ов' if len(paid_ids) > 1 else 'а'
        full_notes = (
            f"Списание долга (обнуление баланса) "
            f"для заказ{order_word} #{', #'.join(paid_ids)}"
            f"\n\nРаспределение:\n" + '\n'.join(distribution)
        )
    else:
        full_notes = 'Списание долга (обнуление баланса)'

    Payment.objects.create(
        client=client,
        amount=debt,
        payment_method='writeoff',
        notes=full_notes,
    )

    log_activity(request.user, 'reset_balance',
                 f'Обнулил баланс клиента {client.get_full_name()} (списано {debt:.0f} сом)')
    messages.success(request, f'Долг клиента {debt:.0f} сом списан. Баланс обнулён.')
    return redirect('main:client_detail', client_id=client_id)


@cashier_required
def apply_credit_to_order(request, order_id):
    """Зачесть аванс клиента в счёт заказа"""
    
    order = get_object_or_404(RentalOrder, id=order_id)
    client = order.client
    
    # Получаем баланс клиента
    credit = client.get_credit()  # Переплата (аванс)
    
    if credit <= 0:
        # Нет аванса для зачёта
        return redirect('main:client_detail', client_id=client.id)
    
    # Получаем долг по заказу
    order_cost = float(order.get_current_total())
    
    # Считаем сколько уже оплачено по этому заказу
    paid_for_order = 0
    for payment in client.payments.all():
        if payment.notes and f'#{order.id}' in payment.notes:
            import re
            for line in payment.notes.split('\n'):
                if f'Заказ #{order.id}:' in line:
                    match = re.search(r'(\d+(?:\.\d+)?)\s*сом', line)
                    if match:
                        paid_for_order += float(match.group(1))
    
    order_debt = order_cost - paid_for_order
    
    if order_debt <= 0:
        # Заказ уже полностью оплачен
        return redirect('main:client_detail', client_id=client.id)
    
    # Зачитываем аванс
    from decimal import Decimal
    
    if credit >= order_debt:
        # Аванса хватает полностью покрыть долг
        amount_to_apply = Decimal(str(order_debt))
    else:
        # Аванса не хватает, зачитываем весь аванс
        amount_to_apply = Decimal(str(credit))
    
    # Создаём оплату как зачёт аванса
    Payment.objects.create(
        client=client,
        amount=amount_to_apply,
        payment_method='credit',  # Зачёт аванса
        notes=f'Зачёт аванса в счёт заказа\n\nРаспределение:\nЗаказ #{order.id}: {amount_to_apply:.0f} сом ({"полностью" if amount_to_apply >= order_debt else "частично"})'
    )
    
    return redirect('main:client_detail', client_id=client.id)

@login_required
def toggle_excluded_day(request, order_id):
    """AJAX: переключить дождливый день для конкретного товара"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)
    order = get_object_or_404(RentalOrder, id=order_id)
    owner = get_tenant_owner(request.user)
    if order.client.owner_id != owner.id:
        return JsonResponse({'error': 'Доступ запрещён'}, status=403)
    try:
        import json as _json
        from datetime import date as date_cls
        from apps.rental.models import OrderExcludedDay
        data = _json.loads(request.body)
        day = date_cls.fromisoformat(data.get('date', ''))
        item_id = int(data.get('item_id', 0))
    except Exception:
        return JsonResponse({'error': 'Неверные данные'}, status=400)
    order_item = get_object_or_404(OrderItem, id=item_id, order=order)
    obj, created = OrderExcludedDay.objects.get_or_create(
        order=order, order_item=order_item, date=day
    )
    if not created:
        obj.delete()
        action = 'removed'
    else:
        action = 'added'
    # Собираем per-item excluded dates
    excluded_by_item = {}
    for item in order.items.all():
        excluded_by_item[str(item.id)] = [
            str(d) for d in item.excluded_days.values_list('date', flat=True)
        ]
    rain_total = order.get_total_excluding_rain()
    original_total = order.get_current_total()
    return JsonResponse({
        'status': action,
        'excluded_by_item': excluded_by_item,
        'rain_total': str(rain_total),
        'original_total': str(original_total),
        'rain_count': order.excluded_days.count(),
    })


@manager_required
def close_order(request, order_id):
    """Закрыть заказ вручную - все товары возвращаются на склад"""
    from apps.rental.models import ReturnDocument, ReturnItem
    from django.utils import timezone
    from decimal import Decimal
    import re

    if request.method != 'POST':
        return redirect('main:client_detail', client_id=get_object_or_404(RentalOrder, id=order_id).client.id)

    order = get_object_or_404(RentalOrder, id=order_id)

    if order.status != 'open':
        messages.warning(request, 'Заказ уже закрыт')
        return redirect('main:client_detail', client_id=order.client.id)

    client = order.client

    # === Проверяем долг по заказу перед закрытием ===
    # Текущая стоимость заказа (с учётом просрочек/изменений)
    current_cost = float(order.get_current_total())

    # Сколько уже оплачено по этому заказу (по распределениям в примечаниях платежей)
    total_paid_for_order = 0.0
    for payment in client.payments.all():
        if payment.notes and f'#{order.id}' in payment.notes:
            if 'Распределение:' in payment.notes:
                for line in payment.notes.split('\n'):
                    pattern = rf'Заказ\s*#{order.id}\s*:\s*(\d+(?:[\.,]\d+)?)\s*сом'
                    match = re.search(pattern, line)
                    if match:
                        total_paid_for_order += float(match.group(1).replace(',', '.'))
            else:
                # Старые платежи могли быть без блока "Распределение"
                # В этом случае считаем, что весь платёж относится к заказу
                total_paid_for_order += float(payment.amount)

    client_balance = client.get_wallet_balance()
    order_debt = float(calculate_order_debt(order, total_paid_for_order, client_balance))

    if order_debt >= 1:  # порог 1 сом защищает от ошибок float-округления
        messages.error(
            request,
            f'Нельзя закрыть заказ, пока есть долг по заказу: {order_debt:.0f} сом. '
            'Сначала примите оплату или зачтите аванс.'
        )
        return redirect('main:client_detail', client_id=client.id)

    now = timezone.now()
    items_to_return = list(order.items.filter(quantity_remaining__gt=0))
    
    if items_to_return:
        return_doc = ReturnDocument.objects.create(
            notes=f'Автоматический возврат при закрытии заказа #{order.id}',
            return_date=now
        )
        
        for item in items_to_return:
            delta = now - item.issued_date
            total_seconds = delta.total_seconds()
            actual_days = int(total_seconds // 86400)
            remaining_seconds = total_seconds % 86400
            actual_hours = int(remaining_seconds // 3600)
            
            calculated_cost = Decimal(item.quantity_remaining) * (
                Decimal(str(item.price_per_day)) * actual_days + 
                Decimal(str(item.price_per_hour)) * actual_hours
            )
            
            ReturnItem.objects.create(
                return_document=return_doc,
                order_item=item,
                quantity=item.quantity_remaining,
                actual_days=actual_days,
                actual_hours=actual_hours,
                calculated_cost=calculated_cost
            )
            
            # Обновляем OrderItem
            item.quantity_returned += item.quantity_remaining
            item.quantity_remaining = 0
            item.actual_cost = calculated_cost
            item.current_total_cost = calculated_cost
            item.save()
            
            # ✅ ОБНОВЛЯЕМ ИНВЕНТАРЬ
            product = item.product
            product.update_available_quantity()
    
    order.status = 'closed'
    order.save()
    log_activity(request.user, 'close_order', f'Закрыл заказ #{order.id} (клиент: {order.client.get_full_name()})')

    # Email клиенту при закрытии
    try:
        from apps.main.email_utils import notify_order_closed_email
        notify_order_closed_email(order)
    except Exception:
        pass

    from apps.main.cache_utils import invalidate_dashboard
    invalidate_dashboard(get_tenant_owner(request.user).id)
    messages.success(request, f'✅ Заказ #{order.id} закрыт. Все товары возвращены на склад.')
    return redirect('main:client_detail', client_id=order.client.id)


from django.db.models.functions import Lower

def global_search(request):
    """Глобальный поиск - фильтрация в Python (работает всегда!)"""

    query = request.GET.get('q', '').strip()

    print(f"Поиск: '{query}'")

    if len(query) < 2:
        return JsonResponse({
            'clients': [],
            'orders': [],
            'products': []
        })

    from apps.clients.models import Client
    from apps.rental.models import RentalOrder
    from apps.inventory.models import Product

    owner = get_tenant_owner(request.user)

    # Приводим к нижнему регистру для поиска
    query_lower = query.lower()

    # ============================================================
    # ПОИСК КЛИЕНТОВ В PYTHON
    # ============================================================

    # Получаем клиентов текущего тенанта
    all_clients = Client.objects.filter(owner=owner).prefetch_related('phones')
    
    found_clients = []
    for client in all_clients:
        # Проверяем в Python (точно работает!)
        match = False
        
        # Проверяем имя
        if client.first_name and query_lower in client.first_name.lower():
            match = True
        
        # Проверяем фамилию
        if client.last_name and query_lower in client.last_name.lower():
            match = True
        
        # Проверяем отчество
        if client.middle_name and query_lower in client.middle_name.lower():
            match = True
        
        # Проверяем телефоны
        for phone in client.phones.all():
            if query in phone.phone_number:
                match = True
                break
        
        if match:
            found_clients.append(client)
            print(f"  ✅ Найден: {client.get_full_name()}")
            
            # Ограничиваем 5 результатами
            if len(found_clients) >= 5:
                break
    
    print(f"📊 Всего найдено клиентов: {len(found_clients)}")
    
    # Формируем результаты
    clients_results = []
    for client in found_clients:
        try:
            balance = float(client.get_wallet_balance())
            badge_color = 'bg-green-100 text-green-700' if balance >= 0 else 'bg-red-100 text-red-700'
            badge_text = f'+{int(balance)}' if balance > 0 else f'{int(balance)}'
            
            phones = ', '.join([p.phone_number for p in client.phones.all()[:2]])
            
            clients_results.append({
                'type': 'client',
                'title': client.get_full_name(),
                'subtitle': phones,
                'badge': f'{badge_text} сом',
                'badge_color': badge_color,
                'url': f'/clients/{client.id}/',
            })
        except Exception as e:
            print(f"❌ Ошибка формирования результата: {e}")
    
    # ============================================================
    # ПОИСК ЗАКАЗОВ В PYTHON
    # ============================================================
    
    found_orders = []
    
    # Если это число - ищем по ID
    try:
        order_id = int(query)
        found_orders = list(RentalOrder.objects.filter(id=order_id, client__owner=owner).select_related('client')[:5])
    except ValueError:
        # Иначе ищем по клиенту
        all_orders = RentalOrder.objects.filter(client__owner=owner).select_related('client').prefetch_related('client__phones')[:100]
        
        for order in all_orders:
            match = False
            
            # Ищем в имени клиента
            if order.client.first_name and query_lower in order.client.first_name.lower():
                match = True
            if order.client.last_name and query_lower in order.client.last_name.lower():
                match = True
            
            # Ищем в примечаниях
            if order.notes and query_lower in order.notes.lower():
                match = True
            
            if match:
                found_orders.append(order)
                if len(found_orders) >= 5:
                    break
    
    orders_results = []
    for order in found_orders:
        try:
            badge_color = 'bg-blue-100 text-blue-700' if order.status == 'open' else 'bg-gray-100 text-gray-700'
            badge_text = 'Открыт' if order.status == 'open' else 'Закрыт'
            
            orders_results.append({
                'type': 'order',
                'title': f'Заказ #{order.id}',
                'subtitle': f'{order.client.get_full_name()} • {order.created_at.strftime("%d.%m.%Y")}',
                'badge': badge_text,
                'badge_color': badge_color,
                'url': f'/clients/{order.client.id}/',
            })
        except Exception as e:
            print(f"❌ Ошибка: {e}")
    
    # ============================================================
    # ПОИСК ТОВАРОВ В PYTHON
    # ============================================================
    
    all_products = Product.objects.filter(owner=owner)
    found_products = []
    
    for product in all_products:
        if product.name and query_lower in product.name.lower():
            found_products.append(product)
            if len(found_products) >= 5:
                break
    
    products_results = []
    for product in found_products:
        try:
            products_results.append({
                'type': 'product',
                'title': product.name,
                'subtitle': f'{product.price_per_day} сом/день',
                'badge': f'{product.quantity_available} шт',
                'badge_color': 'bg-purple-100 text-purple-700',
                'url': f'/admin/inventory/product/{product.id}/change/',
            })
        except Exception as e:
            print(f"❌ Ошибка: {e}")
    
    print(f"📤 Возвращаем: {len(clients_results)} клиентов, {len(orders_results)} заказов, {len(products_results)} товаров")
    
    return JsonResponse({
        'clients': clients_results,
        'orders': orders_results,
        'products': products_results,
    }, json_dumps_params={'ensure_ascii': False})

def global_search_optimized(request):
    """Оптимизированный поиск для больших баз"""
    
    query = request.GET.get('q', '').strip().lower()
    
    if len(query) < 2:
        return JsonResponse({'clients': [], 'orders': [], 'products': []})
    
    from apps.clients.models import Client
    
    # Берём только последних 500 клиентов (самые свежие)
    all_clients = Client.objects.prefetch_related('phones').order_by('-id')[:500]
    
    found_clients = []
    for client in all_clients:
        # Ищем в полном имени сразу (быстрее)
        full_name = f"{client.first_name} {client.last_name} {client.middle_name or ''}".lower()
        
        if query in full_name:
            found_clients.append(client)
            if len(found_clients) >= 5:
                break
    
    
    clients_results = []
    for client in found_clients:
        balance = float(client.get_wallet_balance())
        badge_color = 'bg-green-100 text-green-700' if balance >= 0 else 'bg-red-100 text-red-700'
        badge_text = f'+{int(balance)}' if balance > 0 else f'{int(balance)}'
        phones = ', '.join([p.phone_number for p in client.phones.all()[:2]])
        
        clients_results.append({
            'type': 'client',
            'title': client.get_full_name(),
            'subtitle': phones,
            'badge': f'{badge_text} сом',
            'badge_color': badge_color,
            'url': f'/clients/{client.id}/',
        })
    
    return JsonResponse({
        'clients': clients_results,
        'orders': [],
        'products': [],
    }, json_dumps_params={'ensure_ascii': False})

# ─── SSE: уведомления в реальном времени ─────────────────────────────────────

@login_required
def sse_stream(request):
    """Server-Sent Events: стримит новые уведомления пользователю."""
    import time as _time
    import json as _json

    def event_generator():
        last_id = int(request.GET.get('last_id', 0))
        # Пинг при подключении
        yield "data: {\"type\": \"ping\"}\n\n"
        # Цикл ~2 минуты (30 × 4 сек), затем клиент переподключается сам
        for _ in range(30):
            _time.sleep(4)
            from apps.main.models import Notification
            notes = (Notification.objects
                     .filter(user=request.user, id__gt=last_id, is_read=False)
                     .order_by('id')[:10])
            for note in notes:
                data = _json.dumps({
                    'id': note.id,
                    'type': note.type,
                    'title': note.title,
                    'message': note.message,
                    'link': note.link,
                }, ensure_ascii=False)
                yield f"data: {data}\n\n"
                last_id = note.id

    from django.http import StreamingHttpResponse
    response = StreamingHttpResponse(event_generator(), content_type='text/event-stream')
    response['Cache-Control'] = 'no-cache'
    response['X-Accel-Buffering'] = 'no'
    return response


@login_required
def mark_notifications_read(request):
    """
    GET ?poll=1&last_id=N  — вернуть новые уведомления (для polling).
    POST ids[]             — отметить прочитанными.
    """
    from apps.main.models import Notification
    if request.method == 'POST':
        ids = request.POST.getlist('ids[]')
        if ids:
            Notification.objects.filter(user=request.user, id__in=ids).update(is_read=True)
        count = Notification.objects.filter(user=request.user, is_read=False).count()
        return JsonResponse({'ok': True, 'unread': count})
    # GET — polling
    last_id = int(request.GET.get('last_id', 0))
    notes = list(
        Notification.objects
        .filter(user=request.user, id__gt=last_id)
        .order_by('id')[:20]
        .values('id', 'type', 'title', 'message', 'link', 'is_read', 'created_at')
    )
    for n in notes:
        n['created_at'] = n['created_at'].strftime('%d.%m %H:%M')
    return JsonResponse({'notes': notes})


@login_required
def api_overdue_orders(request):
    """API: список просроченных открытых заказов для колокольчика"""
    now = timezone.now()
    owner = get_tenant_owner(request.user)
    overdue_items = []
    seen_orders = set()

    from apps.rental.models import OrderItem
    items = (OrderItem.objects
             .filter(order__status='open', order__client__owner=owner, quantity_remaining__gt=0, planned_return_date__lt=now)
             .select_related('order', 'order__client', 'product')
             .order_by('planned_return_date'))

    for item in items:
        order = item.order
        if order.id in seen_orders:
            continue
        seen_orders.add(order.id)

        # Collect all overdue products for this order
        overdue_products = (OrderItem.objects
                            .filter(order=order, quantity_remaining__gt=0, planned_return_date__lt=now)
                            .select_related('product'))
        products_str = ', '.join(
            f"{oi.product.name} x{oi.quantity_remaining}"
            for oi in overdue_products
        )
        earliest = overdue_products.order_by('planned_return_date').first()
        since_str = timezone.localtime(earliest.planned_return_date).strftime('%d.%m.%Y %H:%M') if earliest else ''

        overdue_items.append({
            'order_id': order.id,
            'client': order.client.get_full_name(),
            'products': products_str,
            'since': since_str,
            'url': f'/clients/{order.client.id}/',
        })

    return JsonResponse({
        'count': len(overdue_items),
        'orders': overdue_items,
    }, json_dumps_params={'ensure_ascii': False})


def send_overdue_notification(request, order_id):
    '''Ручная отправка уведомления о просрочке'''
    from apps.rental.models import RentalOrder
    from apps.main.telegram_bot_complete import notify_overdue
    from apps.main.email_utils import notify_overdue_email

    order = get_object_or_404(RentalOrder, id=order_id)

    notify_overdue(order)

    email_sent = notify_overdue_email(order)
    if email_sent:
        messages.success(request, f'Уведомление отправлено для заказа #{order_id} (Telegram + Email)')
    else:
        messages.success(request, f'Уведомление отправлено для заказа #{order_id}')

    return redirect('main:client_detail', client_id=order.client.id)



@login_required
def edit_order_dates_with_log(request, order_id):
    """Изменить даты возврата с логированием в notes"""
    order = get_object_or_404(RentalOrder, id=order_id)
    
    if not request.user.is_staff:
        messages.error(request, '🔒 Нет прав')
        return redirect('main:client_detail', client_id=order.client.id)
    
    if request.method == 'POST':
        changes_log = []
        
        for item in order.items.all():
            new_date_str = request.POST.get(f'return_date_{item.id}')
            
            if new_date_str:
                try:
                    new_date = timezone.datetime.strptime(new_date_str, '%Y-%m-%dT%H:%M')
                    new_date = timezone.make_aware(new_date)
                except ValueError:
                    continue
                
                # Сравниваем без секунд (форма не отправляет секунды)
                if new_date == item.planned_return_date.replace(second=0, microsecond=0):
                    continue
                
                old_date = item.planned_return_date
                old_cost = item.original_total_cost
                
                # Пересчёт
                item.planned_return_date = new_date
                item.recalculate_from_dates()
                item.save()
                
                # Логируем изменение
                cost_diff = item.original_total_cost - old_cost
                local_now = timezone.localtime()
                log_entry = (
                    f"[{local_now.strftime('%d.%m.%Y %H:%M')}] "
                    f"{request.user.username}: "
                    f"Изменил дату возврата {item.product.name} x{item.quantity_taken} "
                    f"с {old_date.strftime('%d.%m.%Y %H:%M')} "
                    f"на {new_date.strftime('%d.%m.%Y %H:%M')} "
                    f"({cost_diff:+.0f} сом)"
                )
                changes_log.append(log_entry)
        
        if changes_log:
            # Добавляем в notes заказа
            if order.notes:
                order.notes += '\n\n' + '\n'.join(changes_log)
            else:
                order.notes = '\n'.join(changes_log)
            order.save()
            
            messages.success(request, '✅ Даты обновлены и записаны в историю заказа')
        else:
            messages.info(request, 'ℹ️ Изменений не было')
        
        return redirect('main:client_detail', client_id=order.client.id)
    
    context = {
        'order': order,
        'client': order.client,
    }
    return render(request, 'rental/edit_item_date.html', context)   

@manager_required
def create_product(request):
    '''Создать новый товар'''
    if request.method == 'POST':
        name = request.POST.get('name')
        price_per_day = request.POST.get('price_per_day')
        quantity = request.POST.get('quantity')
        
        product = Product.objects.create(
            name=name,
            price_per_day=price_per_day,
            quantity_available=quantity
        )
        
        messages.success(request, f'Товар "{product.name}" создан!')
        return redirect('/admin/rental/product/')
    
    return render(request, 'rental/create_product.html')

def orders_calendar(request):
    """Календарь заказов"""
    from apps.rental.models import RentalOrder
    from apps.clients.models import Client

    owner = get_tenant_owner(request.user)

    # Получаем выбранного клиента
    selected_client_id = request.GET.get('client_id')
    selected_client_id_int = int(selected_client_id) if selected_client_id else None

    # Базовый запрос - активные заказы текущего тенанта
    orders = RentalOrder.objects.filter(
        status='open', client__owner=owner
    ).select_related('client').prefetch_related('items__product', 'client__phones')

    # Фильтрация по клиенту
    if selected_client_id_int:
        orders = orders.filter(client_id=selected_client_id_int)

    # Получаем список клиентов текущего тенанта для выпадающего списка
    clients = Client.objects.filter(
        owner=owner,
        rental_orders__status='open'
    ).distinct().order_by('last_name', 'first_name')

    # Подсчет активных и просроченных
    now = timezone.now()
    active_count = 0
    overdue_count = 0

    # Формируем события для календаря
    events = []

    for order in orders:
        for item in order.items.filter(quantity_remaining__gt=0):

            # Определяем статус - просрочен или нет
            is_overdue = item.planned_return_date < now

            if is_overdue:
                overdue_count += 1
                status = 'overdue'
            else:
                active_count += 1
                status = 'active'

            # Выдача (pickup) - синий
            events.append({
                'id': f'order_{order.id}_item_{item.id}_pickup',
                'title': f'{item.product.name} x{item.quantity_remaining}',
                'start': item.issued_date.isoformat(),
                'allDay': True,
                'backgroundColor': '#3b82f6',
                'borderColor': '#3b82f6',
                'extendedProps': {
                    'client_name': order.client.get_full_name(),
                    'client_id': order.client.id,
                    'order_id': order.id,
                    'product_name': item.product.name,
                    'quantity': item.quantity_remaining,
                    'phone': ', '.join([p.phone_number for p in order.client.phones.all()]),
                    'status': status,
                    'event_type': 'pickup',
                    'issued_date': item.issued_date.isoformat(),
                    'planned_return_date': item.planned_return_date.isoformat(),
                }
            })

            # Возврат (return) - зеленый или красный если просрочен
            events.append({
                'id': f'order_{order.id}_item_{item.id}_return',
                'title': f'{item.product.name} x{item.quantity_remaining}',
                'start': item.planned_return_date.isoformat(),
                'allDay': True,
                'backgroundColor': '#ef4444' if is_overdue else '#10b981',
                'borderColor': '#ef4444' if is_overdue else '#10b981',
                'extendedProps': {
                    'client_name': order.client.get_full_name(),
                    'client_id': order.client.id,
                    'order_id': order.id,
                    'product_name': item.product.name,
                    'quantity': item.quantity_remaining,
                    'phone': ', '.join([p.phone_number for p in order.client.phones.all()]),
                    'status': status,
                    'event_type': 'return',
                    'issued_date': item.issued_date.isoformat(),
                    'planned_return_date': item.planned_return_date.isoformat(),
                }
            })

    # Дождливые дни этого тенанта
    from apps.main.models import RainDay
    rain_days_qs = RainDay.objects.filter(owner=owner).values_list('date', flat=True)
    rain_days_set = set(str(d) for d in rain_days_qs)

    context = {
        'events_json': json.dumps(events, ensure_ascii=False),
        'rain_days_json': json.dumps(list(rain_days_set), ensure_ascii=False),
        'active_count': active_count,
        'overdue_count': overdue_count,
        'clients': clients,
        'selected_client_id': selected_client_id_int,
    }

    return render(request, 'main/calendar.html', context)


@login_required
def toggle_rain_day(request):
    """AJAX: включить/выключить глобальный дождливый день для тенанта."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)
    try:
        import json as _json
        from datetime import date as date_cls
        from apps.main.models import RainDay
        data = _json.loads(request.body)
        day = date_cls.fromisoformat(data.get('date', ''))
    except Exception:
        return JsonResponse({'error': 'Неверные данные'}, status=400)

    owner = get_tenant_owner(request.user)
    obj, created = RainDay.objects.get_or_create(owner=owner, date=day)
    if not created:
        obj.delete()
        action = 'removed'
    else:
        action = 'added'

    rain_days = list(
        str(d) for d in RainDay.objects.filter(owner=owner).values_list('date', flat=True)
    )
    return JsonResponse({'status': action, 'date': str(day), 'rain_days': rain_days})


@staff_member_required
def download_latest_backup(request):
    """Скачать последний бэкап"""
    
    backup_dir = os.path.join(settings.BASE_DIR, 'backups')
    
    # Находим последний бэкап
    backups = []
    for filename in os.listdir(backup_dir):
        if filename.startswith('db_backup_') and filename.endswith('.sqlite3'):
            filepath = os.path.join(backup_dir, filename)
            backups.append((filepath, os.path.getmtime(filepath)))
    
    if not backups:
        raise Http404('Бэкапы не найдены')
    
    # Последний файл
    backups.sort(key=lambda x: x[1], reverse=True)
    latest_backup = backups[0][0]
    
    # Отдаём файл
    response = FileResponse(open(latest_backup, 'rb'))
    response['Content-Disposition'] = f'attachment; filename="{os.path.basename(latest_backup)}"'
    
    return response


@staff_member_required
def create_backup_now(request):
    """Создать бэкап сейчас"""
    from django.core.management import call_command
    
    try:
        # Вызываем команду создания бэкапа
        call_command('backup_db')
        messages.success(request, '✅ Бэкап успешно создан!')
    except Exception as e:
        messages.error(request, f'❌ Ошибка создания бэкапа: {e}')
    
    return redirect('main:dashboard')

def is_admin(user):
    """Проверка что пользователь администратор"""
    return user.is_superuser or user.groups.filter(name='Администратор').exists()

@admin_required
def users_management(request):
    """Страница управления пользователями и ролями"""

    if request.user.is_staff:
        # Creator sees only directors (is_superuser=True) and himself, NOT employees of directors
        director_ids = UserProfile.objects.filter(
            role='director'
        ).values_list('user_id', flat=True)
        users = User.objects.filter(
            Q(id__in=director_ids) | Q(is_staff=True)
        ).prefetch_related('groups', 'profile').order_by('-date_joined')
    else:
        # Director sees only their employees
        employee_ids = UserProfile.objects.filter(
            owner=request.user
        ).values_list('user_id', flat=True)
        users = User.objects.filter(id__in=employee_ids).prefetch_related('groups', 'profile').order_by('-date_joined')

    # Получаем все группы
    groups = Group.objects.prefetch_related('permissions').all()

    # Статистика по ролям
    stats = {
        'total_users': users.count(),
        'active_users': users.filter(is_active=True).count(),
        'admins': users.filter(Q(is_superuser=True) | Q(groups__name='Администратор')).distinct().count(),
        'managers': users.filter(groups__name='Менеджер').count(),
        'cashiers': users.filter(groups__name='Кассир').count(),
    }

    context = {
        'users': users,
        'groups': groups,
        'stats': stats,
    }

    return render(request, 'main/users_management.html', context)


@login_required
def create_employee(request):
    """Создание сотрудника директором"""
    from django.contrib.auth import get_user_model
    if not request.user.is_superuser:
        messages.error(request, 'Только директор может создавать сотрудников')
        return redirect('main:dashboard')

    UserModel = get_user_model()

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()

        if not username or not password:
            messages.error(request, 'Укажите имя пользователя и пароль')
            return render(request, 'main/create_employee.html', {})

        if UserModel.objects.filter(username=username).exists():
            messages.error(request, f'Пользователь "{username}" уже существует')
            return render(request, 'main/create_employee.html', {})

        employee = UserModel.objects.create_user(
            username=username,
            password=password,
            email=email,
            first_name=first_name,
            last_name=last_name,
            is_active=True,
            is_staff=False,
            is_superuser=False,
        )
        profile, _ = UserProfile.objects.get_or_create(user=employee)
        profile.owner = request.user
        profile.role = 'employee'
        profile.save()
        messages.success(request, f'Сотрудник {username} создан!')
        return redirect('main:users_management')

    return render(request, 'main/create_employee.html', {})


@login_required
@user_passes_test(is_admin)
def toggle_user_active(request, user_id):
    """Активировать/деактивировать пользователя"""
    user = get_object_or_404(User, id=user_id)

    if user == request.user:
        messages.error(request, 'Нельзя деактивировать самого себя!')
        return redirect('main:users_management')

    # Director can only toggle their own employees
    if not request.user.is_staff:
        profile = getattr(user, 'profile', None)
        if not profile or profile.owner != request.user:
            messages.error(request, 'Вы можете управлять только своими сотрудниками!')
            return redirect('main:users_management')

    if user.is_superuser and not request.user.is_staff:
        messages.error(request, 'Нельзя деактивировать суперпользователя!')
        return redirect('main:users_management')

    user.is_active = not user.is_active
    user.save()

    status = 'активирован' if user.is_active else 'деактивирован'
    messages.success(request, f'Пользователь {user.username} {status}')

    referer = request.META.get('HTTP_REFERER', '')
    if 'superadmin' in referer:
        return redirect('main:superuser_panel')
    return redirect('main:users_management')


@login_required
@user_passes_test(is_admin)
def assign_user_group(request, user_id):
    """Назначить группу пользователю"""
    if request.method == 'POST':
        user = get_object_or_404(User, id=user_id)
        group_id = request.POST.get('group_id')
        
        if group_id:
            group = get_object_or_404(Group, id=group_id)
            user.groups.add(group)
            messages.success(request, f'Роль "{group.name}" назначена пользователю {user.username}')
        
        return redirect('main:users_management')


@login_required
@user_passes_test(is_admin)
def remove_user_group(request, user_id, group_id):
    """Удалить группу у пользователя"""
    user = get_object_or_404(User, id=user_id)
    group = get_object_or_404(Group, id=group_id)
    
    user.groups.remove(group)
    messages.success(request, f'Роль "{group.name}" удалена у пользователя {user.username}')
    
    return redirect('main:users_management')


@login_required
def permissions_matrix(request):
    """Матрица прав доступа — редактируемая для директоров"""
    if not (request.user.is_superuser or request.user.groups.filter(name='Администратор').exists()):
        messages.error(request, 'Доступ запрещён')
        return redirect('main:dashboard')

    # Группы, которые директор может назначать сотрудникам
    ASSIGNABLE_GROUPS = ['Менеджер', 'Кассир', 'Администратор']
    groups = Group.objects.filter(name__in=ASSIGNABLE_GROUPS).prefetch_related('permissions')

    # Получаем сотрудников в зависимости от роли пользователя
    if request.user.is_staff:
        # Creator — видит всех (и директоров, и сотрудников)
        employees = User.objects.filter(is_active=True).exclude(id=request.user.id).prefetch_related('groups')
    else:
        # Директор — только свои сотрудники
        emp_ids = UserProfile.objects.filter(owner=request.user).values_list('user_id', flat=True)
        employees = User.objects.filter(id__in=emp_ids, is_active=True).prefetch_related('groups')

    # POST — назначить/снять группу
    if request.method == 'POST':
        target_user_id = request.POST.get('user_id')
        group_name = request.POST.get('group_name')
        action = request.POST.get('action')  # 'add' или 'remove'

        if target_user_id and group_name and action:
            target_user = get_object_or_404(User, id=target_user_id)
            # Проверяем что директор может управлять этим пользователем
            if not request.user.is_staff:
                if not UserProfile.objects.filter(owner=request.user, user=target_user).exists():
                    messages.error(request, 'Нет доступа к этому пользователю')
                    return redirect('main:permissions_matrix')
            # Только разрешённые группы
            if group_name not in ASSIGNABLE_GROUPS:
                messages.error(request, 'Недопустимая роль')
                return redirect('main:permissions_matrix')
            group_obj = get_object_or_404(Group, name=group_name)
            if action == 'add':
                target_user.groups.add(group_obj)
                messages.success(request, f'Роль "{group_name}" назначена {target_user.username}')
            elif action == 'remove':
                target_user.groups.remove(group_obj)
                messages.success(request, f'Роль "{group_name}" снята с {target_user.username}')
        return redirect('main:permissions_matrix')

    context = {
        'groups': groups,
        'employees': employees,
        'is_creator': request.user.is_staff,
    }
    return render(request, 'main/permissions_matrix.html', context)


def superuser_required(view_func):
    """Декоратор — только создатель системы (is_superuser AND is_staff)"""
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('main:login')
        if not (request.user.is_superuser and request.user.is_staff):
            messages.error(request, '⛔ Доступ только для создателя системы')
            return redirect('main:dashboard')
        return view_func(request, *args, **kwargs)
    wrapper.__name__ = view_func.__name__
    return wrapper


@superuser_required
def superuser_panel(request):
    """Панель суперадминистратора — только для создателя системы"""
    from apps.rental.models import RentalOrder
    from apps.clients.models import Client

    # Пользователи, ожидающие одобрения — только Director-регистрации (не сотрудники)
    pending_users = User.objects.filter(
        is_active=False, profile__owner__isnull=True
    ).order_by('date_joined')

    # Одобрение пользователя через POST
    if request.method == 'POST':
        action = request.POST.get('action')
        uid = request.POST.get('user_id')
        target = get_object_or_404(User, id=uid)
        if action == 'approve':
            # Новый пользователь становится суперадмином своей компании
            approved_user = target
            approved_user.is_active = True
            approved_user.is_staff = False   # не системный создатель
            approved_user.is_superuser = True  # владелец своей компании
            approved_user.save()
            admin_group, _ = Group.objects.get_or_create(name='Администратор')
            approved_user.groups.add(admin_group)
            # Профиль с ролью director
            from apps.inventory.models import Warehouse
            profile, _ = UserProfile.objects.get_or_create(user=approved_user, defaults={'owner': None})
            profile.role = 'director'
            profile.needs_company_setup = True
            profile.save()
            Warehouse.objects.get_or_create(
                owner=approved_user,
                name='Основной склад',
                defaults={'description': 'Склад по умолчанию'},
            )
            messages.success(request, f'✅ Пользователь «{approved_user.username}» одобрен. Теперь у него своя пустая CRM.')
        elif action == 'reject':
            target.delete()
            messages.success(request, f'🗑 Пользователь «{target.username}» удалён.')
        return redirect('main:superuser_panel')

    # Superuser panel always shows system-wide totals (all tenants)
    stats = {
        'total_users':   User.objects.count(),
        'pending_users': pending_users.count(),
        'total_clients': Client.objects.count(),   # system-wide
        'open_orders':   RentalOrder.objects.filter(status='open').count(),  # system-wide
    }

    # Director list (merged from creator_directors)
    from apps.inventory.models import Product, Warehouse
    from apps.rental.models import Payment as _Pmt
    from django.db.models import Sum as _Sum
    directors = User.objects.filter(is_superuser=True, is_staff=False, is_active=True)
    director_data = []
    for d in directors:
        from apps.clients.models import Client as _Cli
        _clients = _Cli.objects.filter(owner=d).count()
        _products = Product.objects.filter(owner=d).count()
        _orders = RentalOrder.objects.filter(client__owner=d, status='open').count()
        _revenue = _Pmt.objects.filter(client__owner=d).aggregate(t=_Sum('amount'))['t'] or 0
        _employees = UserProfile.objects.filter(owner=d).count()
        try:
            _max_wh = d.profile.max_warehouses
        except Exception:
            _max_wh = 1
        director_data.append({
            'user': d,
            'clients': _clients,
            'products': _products,
            'open_orders': _orders,
            'revenue': int(_revenue),
            'max_warehouses': _max_wh,
            'employees': _employees,
        })

    # All users for the users tab (same logic as users_management for creator)
    director_ids = UserProfile.objects.filter(role='director').values_list('user_id', flat=True)
    all_users_list = User.objects.filter(
        Q(id__in=director_ids) | Q(is_staff=True)
    ).prefetch_related('groups', 'profile').order_by('-date_joined')
    groups = Group.objects.all()

    return render(request, 'main/superuser_panel.html', {
        'pending_users': pending_users,
        'stats': stats,
        'all_users': all_users_list,
        'director_data': director_data,
        'groups': groups,
    })


@superuser_required
def creator_directors(request):
    """Список всех директоров (для создателя)"""
    from apps.inventory.models import Product, Warehouse
    from apps.clients.models import Client
    from apps.rental.models import RentalOrder, Payment
    from django.db.models import Sum
    import csv
    from django.http import HttpResponse

    # Export CSV
    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="directors.csv"'
        writer = csv.writer(response)
        writer.writerow(['Директор', 'Email', 'Дата регистрации', 'Клиентов', 'Товаров', 'Открытых заказов', 'Доход (сом)', 'Макс складов'])
        for d in User.objects.filter(is_superuser=True, is_staff=False, is_active=True):
            clients = Client.objects.filter(owner=d).count()
            products = Product.objects.filter(owner=d).count()
            orders = RentalOrder.objects.filter(client__owner=d, status='open').count()
            revenue = Payment.objects.filter(client__owner=d).aggregate(t=Sum('amount'))['t'] or 0
            try:
                max_wh = d.profile.max_warehouses
            except Exception:
                max_wh = 1
            writer.writerow([d.username, d.email, d.date_joined.strftime('%d.%m.%Y'), clients, products, orders, int(revenue), max_wh])
        return response

    directors = User.objects.filter(is_superuser=True, is_staff=False, is_active=True)
    director_data = []
    for d in directors:
        clients = Client.objects.filter(owner=d).count()
        products = Product.objects.filter(owner=d).count()
        orders = RentalOrder.objects.filter(client__owner=d, status='open').count()
        revenue = Payment.objects.filter(client__owner=d).aggregate(t=Sum('amount'))['t'] or 0
        try:
            max_wh = d.profile.max_warehouses
        except Exception:
            max_wh = 1
        warehouses = list(Warehouse.objects.filter(owner=d).values_list('name', flat=True))
        employees = UserProfile.objects.filter(owner=d).count()
        director_data.append({
            'user': d,
            'clients': clients,
            'products': products,
            'open_orders': orders,
            'revenue': int(revenue),
            'max_warehouses': max_wh,
            'warehouses': warehouses,
            'employees': employees,
        })

    return render(request, 'main/creator_directors.html', {'directors': director_data})


@superuser_required
def edit_director_settings(request, user_id):
    """Редактировать настройки директора (макс. склады)"""
    target = get_object_or_404(User, id=user_id, is_superuser=True, is_staff=False)
    profile, _ = UserProfile.objects.get_or_create(user=target)

    if request.method == 'POST':
        try:
            profile.max_warehouses = int(request.POST.get('max_warehouses', 1))
            profile.save()
            messages.success(request, f'Настройки {target.username} обновлены')
        except ValueError:
            messages.error(request, 'Некорректное значение')
        return redirect('main:creator_directors')

    return render(request, 'main/edit_director_settings.html', {'target': target, 'profile': profile})


@superuser_required
def creator_director_detail(request, user_id):
    """Детальная страница директора — его сотрудники и склады"""
    from apps.inventory.models import Warehouse
    from apps.clients.models import Client
    from apps.rental.models import RentalOrder, Payment
    from apps.main.models import Expense
    from django.db.models import Sum
    from django.utils import timezone

    director = get_object_or_404(User, id=user_id, is_superuser=True, is_staff=False)
    employees = UserProfile.objects.filter(owner=director).select_related('user').prefetch_related('user__groups')
    warehouses = Warehouse.objects.filter(owner=director)

    # Клиенты
    clients = Client.objects.filter(owner=director).prefetch_related('phones').order_by('-created_at')

    # Финансы
    now = timezone.now()
    total_revenue = Payment.objects.filter(client__owner=director).aggregate(t=Sum('amount'))['t'] or 0
    total_expenses = Expense.objects.filter(owner=director).aggregate(t=Sum('amount'))['t'] or 0
    profit = float(total_revenue) - float(total_expenses)

    # Заказы
    open_orders = RentalOrder.objects.filter(client__owner=director, status='open').count()
    closed_orders = RentalOrder.objects.filter(client__owner=director, status='closed').count()

    # Доходы по месяцам (последние 6)
    from datetime import timedelta
    monthly = []
    for i in range(5, -1, -1):
        month_start = (now.replace(day=1) - timedelta(days=i*28)).replace(day=1)
        if i > 0:
            next_month = (month_start + timedelta(days=32)).replace(day=1)
        else:
            next_month = now
        rev = Payment.objects.filter(
            client__owner=director,
            payment_date__gte=month_start,
            payment_date__lt=next_month,
        ).aggregate(t=Sum('amount'))['t'] or 0
        monthly.append({'label': month_start.strftime('%b'), 'revenue': int(rev)})

    return render(request, 'main/creator_director_detail.html', {
        'director': director,
        'employees': employees,
        'warehouses': warehouses,
        'clients': clients,
        'total_revenue': int(total_revenue),
        'total_expenses': int(total_expenses),
        'profit': int(profit),
        'open_orders': open_orders,
        'closed_orders': closed_orders,
        'monthly': monthly,
    })


@login_required
def ticket_list(request):
    """Список обращений — для директора его тикеты, для создателя все."""
    from apps.main.models import DirectorMessage
    from django.utils import timezone

    if request.user.is_staff:
        tickets = DirectorMessage.objects.select_related('sender').all()
    else:
        tickets = DirectorMessage.objects.filter(sender=request.user)
        # Отметить ответы создателя как прочитанные
        tickets.exclude(reply='').filter(reply_read=False).update(reply_read=True)

    page_obj, page_query = paginate(request, tickets, per_page=20)

    return render(request, 'main/ticket_list.html', {
        'tickets': page_obj,
        'page_obj': page_obj,
        'page_query': page_query,
    })


@login_required
def ticket_detail(request, ticket_id):
    """Просмотр тикета — чат между директором и создателем."""
    from apps.main.models import DirectorMessage, TicketReply

    if request.user.is_staff:
        ticket = get_object_or_404(DirectorMessage, id=ticket_id)
        if not ticket.is_read:
            ticket.is_read = True
            ticket.save(update_fields=['is_read'])
        # Отметить все сообщения от директора как прочитанные
        ticket.replies.filter(is_read=False, author=ticket.sender).update(is_read=True)
    else:
        ticket = get_object_or_404(DirectorMessage, id=ticket_id, sender=request.user)
        # Отметить все ответы создателя как прочитанные
        ticket.replies.filter(is_read=False).exclude(author=request.user).update(is_read=True)

    replies = ticket.replies.select_related('author').all()
    return render(request, 'main/ticket_detail.html', {'ticket': ticket, 'replies': replies})


@login_required
def add_reply(request, ticket_id):
    """Добавить сообщение в чат тикета (и директор, и создатель)."""
    from apps.main.models import DirectorMessage, TicketReply

    if request.user.is_staff:
        ticket = get_object_or_404(DirectorMessage, id=ticket_id)
    else:
        ticket = get_object_or_404(DirectorMessage, id=ticket_id, sender=request.user)

    if not ticket.is_open:
        messages.error(request, 'Обращение закрыто')
        return redirect('main:ticket_detail', ticket_id=ticket.id)

    if request.method == 'POST':
        text = request.POST.get('text', '').strip()
        if text:
            TicketReply.objects.create(ticket=ticket, author=request.user, text=text)
            # Сбросить счётчик прочтения для другой стороны
            if not request.user.is_staff:
                ticket.is_read = False
                ticket.save(update_fields=['is_read'])
            # Push-уведомление другой стороне
            from apps.main.notification_utils import push_notification
            if request.user.is_staff:
                # Создатель ответил → уведомить директора
                push_notification(ticket.sender_id, f'💬 Ответ по тикету #{ticket.pk}',
                                  f'{request.user.username}: {text[:80]}',
                                  type='ticket', link=f'/messages/{ticket.id}/')
            else:
                # Директор написал → уведомить создателя (is_staff)
                from django.contrib.auth.models import User as _User
                creator = _User.objects.filter(is_staff=True, is_superuser=True).first()
                if creator:
                    push_notification(creator.id, f'💬 Новое сообщение в тикете #{ticket.pk}',
                                      f'{request.user.username}: {text[:80]}',
                                      type='ticket', link=f'/messages/{ticket.id}/')
        else:
            messages.error(request, 'Введите текст')

    return redirect('main:ticket_detail', ticket_id=ticket.id)


@login_required
def send_message(request):
    """Создать новое обращение (директор/сотрудник)."""
    from apps.main.models import DirectorMessage, TicketReply

    if request.user.is_staff:
        return redirect('main:ticket_list')

    if request.method == 'POST':
        subject = request.POST.get('subject', '').strip()
        message_text = request.POST.get('message', '').strip()
        if subject and message_text:
            ticket = DirectorMessage.objects.create(
                sender=request.user,
                subject=subject,
                message=message_text,
            )
            # Первое сообщение добавляем как reply тоже
            TicketReply.objects.create(ticket=ticket, author=request.user, text=message_text, is_read=True)
            # Уведомить создателя системы
            creator = User.objects.filter(is_staff=True, is_superuser=True).first()
            if creator:
                from apps.main.notification_utils import push_notification
                push_notification(creator.id,
                                  f'💬 Новое обращение {ticket.ticket_number}',
                                  f'{request.user.username}: {subject}',
                                  type='ticket', link=f'/messages/{ticket.id}/')
            messages.success(request, f'✅ Обращение {ticket.ticket_number} создано!')
            return redirect('main:ticket_detail', ticket_id=ticket.id)
        messages.error(request, 'Заполните тему и сообщение')

    return render(request, 'main/send_message.html', {})


@login_required
def edit_ticket(request, ticket_id):
    """Директор редактирует первое сообщение (только если нет ответов)."""
    from apps.main.models import DirectorMessage

    ticket = get_object_or_404(DirectorMessage, id=ticket_id, sender=request.user)
    if not ticket.is_open or ticket.replies.exclude(author=request.user).exists():
        messages.error(request, 'Редактирование недоступно после ответа администратора')
        return redirect('main:ticket_detail', ticket_id=ticket.id)

    if request.method == 'POST':
        subject = request.POST.get('subject', '').strip()
        message_text = request.POST.get('message', '').strip()
        if subject and message_text:
            ticket.subject = subject
            ticket.message = message_text
            ticket.save(update_fields=['subject', 'message', 'updated_at'])
            # Обновить текст первого reply
            first_reply = ticket.replies.filter(author=request.user).first()
            if first_reply:
                first_reply.text = message_text
                first_reply.save(update_fields=['text'])
            messages.success(request, '✅ Обращение обновлено')
            return redirect('main:ticket_detail', ticket_id=ticket.id)
        messages.error(request, 'Заполните поля')

    return render(request, 'main/send_message.html', {'ticket': ticket, 'editing': True})


@superuser_required
def reply_ticket(request, ticket_id):
    """Legacy redirect — теперь используется add_reply."""
    return redirect('main:ticket_detail', ticket_id=ticket_id)


@superuser_required
def close_ticket(request, ticket_id):
    """Создатель закрывает обращение."""
    from apps.main.models import DirectorMessage

    if request.method == 'POST':
        DirectorMessage.objects.filter(id=ticket_id).update(status='closed')
    return redirect('main:ticket_detail', ticket_id=ticket_id)


@superuser_required
def mark_message_read(request, msg_id):
    """Пометить сообщение как прочитанное (legacy)."""
    from apps.main.models import DirectorMessage
    DirectorMessage.objects.filter(id=msg_id).update(is_read=True)
    return redirect('main:ticket_list')


# ─── Смена пароля ────────────────────────────────────────────────────────────

@login_required
def change_user_password(request, user_id):
    """
    Создатель меняет пароль директорам.
    Директор меняет пароль своим сотрудникам.
    """
    target = get_object_or_404(User, id=user_id)

    # Определяем доступ
    if request.user.is_staff and request.user.is_superuser:
        # Создатель может менять любому
        pass
    elif request.user.is_superuser and not request.user.is_staff:
        # Директор — только своим сотрудникам
        if not UserProfile.objects.filter(owner=request.user, user=target).exists():
            messages.error(request, 'Нет доступа к этому пользователю')
            return redirect('main:users_management')
    else:
        messages.error(request, 'Доступ запрещён')
        return redirect('main:dashboard')

    if request.method == 'POST':
        new_password = request.POST.get('new_password', '').strip()
        confirm = request.POST.get('confirm_password', '').strip()
        if not new_password:
            messages.error(request, 'Введите новый пароль')
        elif new_password != confirm:
            messages.error(request, 'Пароли не совпадают')
        elif len(new_password) < 6:
            messages.error(request, 'Пароль должен быть не менее 6 символов')
        else:
            target.set_password(new_password)
            target.save()
            messages.success(request, f'Пароль пользователя {target.username} изменён')
            return redirect('main:users_management')

    return render(request, 'main/change_password.html', {'target_user': target})


# ─── Лог активности сотрудника ───────────────────────────────────────────────

@login_required
def employee_activity(request, user_id):
    """Лог действий сотрудника — для директора"""
    from apps.main.models import ActivityLog

    target = get_object_or_404(User, id=user_id)

    # Доступ: создатель или директор этого сотрудника
    if request.user.is_staff and request.user.is_superuser:
        pass  # всё разрешено
    elif request.user.is_superuser and not request.user.is_staff:
        if not UserProfile.objects.filter(owner=request.user, user=target).exists():
            # Также разрешаем смотреть на самого директора (свой лог)
            if target != request.user:
                messages.error(request, 'Нет доступа')
                return redirect('main:users_management')
    else:
        messages.error(request, 'Доступ запрещён')
        return redirect('main:dashboard')

    logs = ActivityLog.objects.filter(user=target).order_by('-created_at')[:200]

    context = {
        'target_user': target,
        'logs': logs,
    }
    return render(request, 'main/employee_activity.html', context)


@login_required
def audit_log(request):
    """Полный аудит-лог всех действий компании."""
    from apps.main.models import ActivityLog

    owner = get_tenant_owner(request.user)

    # Все пользователи тенанта
    if request.user.is_staff:
        user_ids = list(User.objects.values_list('id', flat=True))
    else:
        emp_ids = list(UserProfile.objects.filter(owner=owner).values_list('user_id', flat=True))
        user_ids = emp_ids + [owner.id]

    logs = ActivityLog.objects.filter(user_id__in=user_ids).select_related('user').order_by('-created_at')

    # Фильтры
    action_filter = request.GET.get('action', '')
    user_filter = request.GET.get('user', '')
    date_from = request.GET.get('from', '')
    date_to = request.GET.get('to', '')

    if action_filter:
        logs = logs.filter(action=action_filter)
    if user_filter:
        logs = logs.filter(user_id=user_filter)
    if date_from:
        logs = logs.filter(created_at__date__gte=date_from)
    if date_to:
        logs = logs.filter(created_at__date__lte=date_to)

    page_obj, page_query = paginate(request, logs, per_page=50)

    users = User.objects.filter(id__in=user_ids).order_by('username')
    actions = ActivityLog.ACTION_CHOICES

    return render(request, 'main/audit_log.html', {
        'logs': page_obj,
        'page_obj': page_obj,
        'page_query': page_query,
        'users': users,
        'actions': actions,
        'action_filter': action_filter,
        'user_filter': user_filter,
        'date_from': date_from,
        'date_to': date_to,
    })


@login_required
def order_view(request, order_id):
    """Просмотр заказа — всё на одной странице: товары, дождевой календарь, возвраты, заметка"""
    from apps.rental.models import ReturnItem
    from datetime import date as date_cls, timedelta as td
    import json as _json

    owner = get_tenant_owner(request.user)
    order = get_object_or_404(RentalOrder, id=order_id, client__owner=owner)
    client = order.client

    # Сохранение заметки
    if request.method == 'POST':
        order.notes = request.POST.get('notes', '')
        order.save(update_fields=['notes'])
        messages.success(request, 'Заметка сохранена')
        return redirect('main:order_view', order_id=order.id)

    items = list(order.items.select_related('product').prefetch_related('excluded_days').all())

    # Строим диапазон дат для календаря
    cal_start = cal_end = None
    for it in items:
        if it.issued_date:
            d = it.issued_date.date() if hasattr(it.issued_date, 'date') else it.issued_date
            cal_start = d if cal_start is None else min(cal_start, d)
        if it.planned_return_date:
            d = it.planned_return_date.date() if hasattr(it.planned_return_date, 'date') else it.planned_return_date
            cal_end = d if cal_end is None else max(cal_end, d)

    calendar_days = []
    if cal_start and cal_end:
        cur = cal_start
        while cur <= cal_end and (cur - cal_start).days < 91:
            calendar_days.append(cur)
            cur += td(days=1)

    # Per-item excluded dates + item date ranges для отображения в шаблоне
    item_rain_data = []
    for it in items:
        excl = set(str(d) for d in it.excluded_days.values_list('date', flat=True))
        it_start = it.issued_date.date() if it.issued_date and hasattr(it.issued_date, 'date') else it.issued_date
        it_end = it.planned_return_date.date() if it.planned_return_date and hasattr(it.planned_return_date, 'date') else it.planned_return_date
        item_rain_data.append({
            'item': it,
            'excluded': list(excl),
            'start': str(it_start) if it_start else None,
            'end': str(it_end) if it_end else None,
        })

    excluded_by_item_json = _json.dumps({
        str(row['item'].id): row['excluded'] for row in item_rain_data
    })

    rain_total = order.get_total_excluding_rain()
    original_total = order.get_current_total()

    # Разбивка стоимости: базовая аренда + просрочка + ремонт
    from decimal import Decimal as _D
    base_total = order.get_original_total()

    # Считаем ремонт/чистку отдельно от просрочки
    repair_items_qs = ReturnItem.objects.filter(
        order_item__order=order, repair_fee__gt=0
    ).values_list('repair_fee', 'repair_notes')
    total_repair_fee = sum(float(r[0]) for r in repair_items_qs)
    repair_details = [{'fee': float(r[0]), 'notes': r[1] or 'Ремонт/чистка'} for r in repair_items_qs]

    overdue_total = original_total - base_total - _D(str(total_repair_fee))

    # Детальный разбор штрафа (по каждому товару)
    _now = timezone.now()
    overdue_breakdown = []
    for _item in order.items.all():
        # Возвращённые поздно
        for _ret in _item.returns.select_related('return_document').all():
            if _ret.actual_days > _item.rental_days or (_ret.actual_days == _item.rental_days and _ret.actual_hours > 0 and _item.price_per_hour == 0):
                _planned = float(_item.price_per_day) * _item.rental_days * _ret.quantity
                _actual_base = float(_ret.calculated_cost) - float(_ret.repair_fee or 0)
                _extra = round(_actual_base - _planned, 2)
                if _extra > 0:
                    overdue_breakdown.append({
                        'type': 'late_return',
                        'product': _item.product.name,
                        'quantity': _ret.quantity,
                        'planned_days': _item.rental_days,
                        'actual_days': _ret.actual_days,
                        'actual_hours': _ret.actual_hours,
                        'extra_days': _ret.actual_days - _item.rental_days,
                        'price_per_day': float(_item.price_per_day),
                        'extra_cost': _extra,
                        'date': timezone.localtime(_ret.return_document.return_date).strftime('%d.%m.%Y'),
                    })
        # Ещё не вернули, просрочка
        if _item.quantity_remaining > 0 and _item.planned_return_date < _now:
            _delta = _now - _item.planned_return_date
            _od = _delta.days
            _oh = _delta.seconds // 3600
            _oc = round(float(_item.price_per_day) * _od * _item.quantity_remaining, 2)
            overdue_breakdown.append({
                'type': 'still_out',
                'product': _item.product.name,
                'quantity': _item.quantity_remaining,
                'overdue_days': _od,
                'overdue_hours': _oh,
                'price_per_day': float(_item.price_per_day),
                'overdue_cost': _oc,
                'planned_return': timezone.localtime(_item.planned_return_date).strftime('%d.%m.%Y'),
            })

    # Глобальные дождливые дни тенанта
    from apps.main.models import RainDay
    global_rain_days = set(
        str(d) for d in RainDay.objects.filter(owner=owner).values_list('date', flat=True)
    )
    order_rain_days_json = _json.dumps(list(global_rain_days))

    # История возвратов
    returns = ReturnItem.objects.filter(
        order_item__order=order
    ).select_related('return_document', 'order_item__product').order_by('-return_document__return_date')

    # Оплаты по этому заказу
    import re as _re_ov
    order_payments_list = []
    total_paid_order = 0.0
    for p in client.payments.all().order_by('-payment_date'):
        notes_text = p.notes or ''
        attributed = 0.0
        if 'Распределение:' in notes_text:
            for line in notes_text.split('\n'):
                m = _re_ov.search(rf'Заказ\s*#{order.id}\s*:\s*(\d+(?:[\.,]\d+)?)\s*сом', line)
                if m:
                    attributed = float(m.group(1).replace(',', '.'))
                    break
        elif f'Заказ #{order.id}' in notes_text or f'#{order.id}' in notes_text:
            attributed = float(p.amount)
        if attributed > 0:
            order_payments_list.append({'payment': p, 'amount': attributed})
            total_paid_order += attributed

    final_total_val = float(rain_total) if order.get_rain_excluded_count() > 0 else float(original_total)
    # Прибавляем стоимость доставки (если > 0)
    if order.has_delivery and order.delivery_cost > 0:
        final_total_val += float(order.delivery_cost)
    total_owed_order = max(0.0, final_total_val - total_paid_order)

    return render(request, 'rental/order_view.html', {
        'order': order,
        'client': client,
        'items': items,
        'item_rain_data': item_rain_data,
        'calendar_days': calendar_days,
        'excluded_by_item_json': excluded_by_item_json,
        'rain_excluded_count': order.get_rain_excluded_count(),
        'rain_total': rain_total,
        'original_total': original_total,
        'base_total': base_total,
        'overdue_total': overdue_total,
        'overdue_breakdown': overdue_breakdown,
        'total_repair_fee': total_repair_fee,
        'repair_details': repair_details,
        'global_rain_days_json': order_rain_days_json,
        'returns': returns,
        'now': timezone.now(),
        'has_overdue': any(it.is_overdue for it in items),
        'order_payments_list': order_payments_list,
        'total_paid_order': total_paid_order,
        'total_owed_order': total_owed_order,
        'final_total_val': final_total_val,
        'attachments': order.attachments.select_related('uploaded_by').all(),
    })


@login_required
def broadcast_notifications(request):
    """Массовая рассылка уведомлений клиентам через Telegram"""
    from apps.rental.models import RentalOrder
    from apps.main.models import UserProfile
    from apps.main.telegram_bot_complete import notify_overdue, notify_debt_reminder, send_custom_broadcast

    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, '🔒 Нет прав доступа')
        return redirect('main:dashboard')

    is_creator = request.user.is_staff
    owner = get_tenant_owner(request.user)

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'overdue':
            sent, skipped = 0, 0
            seen_clients = set()
            for order in RentalOrder.objects.filter(
                client__owner=owner, status='open'
            ).select_related('client').prefetch_related('items'):
                if order.client_id in seen_clients:
                    continue
                if any(it.is_overdue for it in order.items.all()):
                    ok = notify_overdue(order)
                    seen_clients.add(order.client_id)
                    if ok:
                        sent += 1
                    else:
                        skipped += 1
            messages.success(request, f'✅ Отправлено {sent} уведомлений о просрочке. Без Telegram: {skipped}')

        elif action == 'debt':
            sent, skipped = 0, 0
            for client in Client.objects.filter(owner=owner):
                ok = notify_debt_reminder(client)
                if ok:
                    sent += 1
                elif client.telegram_id and client.has_debt():
                    skipped += 1
            messages.success(request, f'✅ Отправлено {sent} напоминаний о долге. Без Telegram: {skipped}')

        elif action == 'custom':
            custom_msg = request.POST.get('custom_message', '').strip()
            target = request.POST.get('target', 'all')
            if not custom_msg:
                messages.error(request, '⚠️ Введите текст сообщения')
                return redirect('main:broadcast_notifications')

            all_clients = Client.objects.filter(owner=owner)

            if target == 'overdue':
                overdue_ids = set()
                for order in RentalOrder.objects.filter(
                    client__owner=owner, status='open'
                ).prefetch_related('items'):
                    if any(it.is_overdue for it in order.items.all()):
                        overdue_ids.add(order.client_id)
                chat_ids = list(
                    all_clients.filter(id__in=overdue_ids)
                    .exclude(telegram_id__isnull=True).exclude(telegram_id='')
                    .values_list('telegram_id', flat=True)
                )
            elif target == 'debtors':
                chat_ids = [c.telegram_id for c in all_clients if c.has_debt() and c.telegram_id]
            else:  # all
                chat_ids = list(
                    all_clients.exclude(telegram_id__isnull=True).exclude(telegram_id='')
                    .values_list('telegram_id', flat=True)
                )

            sent, failed = send_custom_broadcast(chat_ids, custom_msg)
            messages.success(request, f'✅ Отправлено {sent} сообщений клиентам. Ошибок: {failed}')

        elif action == 'directors' and is_creator:
            custom_msg = request.POST.get('directors_message', '').strip()
            if not custom_msg:
                messages.error(request, '⚠️ Введите текст сообщения для директоров')
                return redirect('main:broadcast_notifications')
            profiles = UserProfile.objects.filter(role='director').exclude(telegram_chat_id='')
            chat_ids = list(profiles.values_list('telegram_chat_id', flat=True))
            sent, failed = send_custom_broadcast(chat_ids, custom_msg)
            messages.success(request, f'✅ Отправлено {sent} сообщений директорам. Ошибок: {failed}')

        elif action == 'employees' and not is_creator:
            custom_msg = request.POST.get('employees_message', '').strip()
            if not custom_msg:
                messages.error(request, '⚠️ Введите текст сообщения для сотрудников')
                return redirect('main:broadcast_notifications')
            profiles = UserProfile.objects.filter(owner=request.user).exclude(telegram_chat_id='')
            chat_ids = list(profiles.values_list('telegram_chat_id', flat=True))
            sent, failed = send_custom_broadcast(chat_ids, custom_msg)
            messages.success(request, f'✅ Отправлено {sent} сообщений сотрудникам. Ошибок: {failed}')

        return redirect('main:broadcast_notifications')

    # ── GET: статистика ──
    all_clients = Client.objects.filter(owner=owner)
    total_clients = all_clients.count()
    with_telegram = all_clients.exclude(telegram_id__isnull=True).exclude(telegram_id='').count()

    overdue_client_ids = set()
    for order in RentalOrder.objects.filter(
        client__owner=owner, status='open'
    ).prefetch_related('items'):
        if any(it.is_overdue for it in order.items.all()):
            overdue_client_ids.add(order.client_id)
    overdue_count = len(overdue_client_ids)
    debtor_count = sum(1 for c in all_clients if c.has_debt())

    # Статистика для создателя — директора с Telegram
    directors_with_tg = 0
    if is_creator:
        directors_with_tg = UserProfile.objects.filter(role='director').exclude(telegram_chat_id='').count()

    # Статистика для директора — сотрудники с Telegram
    employees_with_tg = 0
    if not is_creator:
        employees_with_tg = UserProfile.objects.filter(owner=request.user).exclude(telegram_chat_id='').count()

    return render(request, 'main/broadcast_notifications.html', {
        'total_clients': total_clients,
        'with_telegram': with_telegram,
        'overdue_count': overdue_count,
        'debtor_count': debtor_count,
        'is_creator': is_creator,
        'directors_with_tg': directors_with_tg,
        'employees_with_tg': employees_with_tg,
    })


# ─── Вспомогательная функция логирования ────────────────────────────────────

def log_activity(user, action, description):
    """Записывает действие пользователя в ActivityLog"""
    try:
        from apps.main.models import ActivityLog
        ActivityLog.objects.create(user=user, action=action, description=description)
    except Exception:
        pass  # Логирование не должно ломать основную логику



@login_required
def my_profile(request):
    from apps.main.models import UserProfile
    profile, _ = UserProfile.objects.get_or_create(user=request.user)

    if request.method == 'POST':
        telegram_chat_id = request.POST.get('telegram_chat_id', '').strip()
        if telegram_chat_id and not telegram_chat_id.lstrip('-').isdigit():
            messages.error(request, 'Telegram ID должен быть числом (например: 123456789)')
        else:
            existing = UserProfile.objects.filter(
                telegram_chat_id=telegram_chat_id
            ).exclude(user=request.user).first() if telegram_chat_id else None
            if existing:
                messages.error(request, 'Этот Telegram ID уже привязан к другому аккаунту')
            else:
                profile.telegram_chat_id = telegram_chat_id
                profile.save()
                messages.success(request, '✅ Telegram привязан!' if telegram_chat_id else 'Telegram отвязан')
        return redirect('main:my_profile')

    return render(request, 'main/my_profile.html', {'profile': profile})


def custom_404_view(request, exception=None):
    """Кастомный обработчик 404 — редирект на главную с сообщением"""
    messages.error(request, '🔍 Страница не найдена. Возможно, она была удалена или URL введён неверно.')
    return redirect('main:dashboard')


# ─── Расходы ───────────────────────────────────────────────────────────────

@login_required
def expenses_list(request):
    from apps.main.models import Expense
    from django.db.models import Sum
    owner = get_tenant_owner(request.user)

    # Фильтр по месяцу/году
    now = timezone.now()
    try:
        year  = int(request.GET.get('year', now.year))
        month = int(request.GET.get('month', now.month))
    except (ValueError, TypeError):
        year, month = now.year, now.month

    expenses = Expense.objects.filter(owner=owner, date__year=year, date__month=month)
    total    = expenses.aggregate(t=Sum('amount'))['t'] or 0

    # По категориям
    by_category = {}
    for exp in expenses:
        cat = exp.get_category_display()
        by_category[cat] = by_category.get(cat, 0) + float(exp.amount)

    # Месяцы для навигации (текущий год)
    months = [
        {'num': m, 'name': ['Январь','Февраль','Март','Апрель','Май','Июнь',
                             'Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь'][m-1]}
        for m in range(1, 13)
    ]

    categories = Expense.CATEGORY_CHOICES

    page_obj, page_query = paginate(request, expenses, per_page=30)

    return render(request, 'main/expenses.html', {
        'expenses': page_obj,
        'page_obj': page_obj,
        'page_query': page_query,
        'total': total,
        'by_category': by_category,
        'year': year,
        'month': month,
        'months': months,
        'categories': categories,
        'month_name': ['Январь','Февраль','Март','Апрель','Май','Июнь',
                       'Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь'][month-1],
    })


@login_required
def create_expense(request):
    from apps.main.models import Expense
    if request.method != 'POST':
        return redirect('main:expenses_list')

    owner = get_tenant_owner(request.user)
    try:
        amount = float(request.POST.get('amount', 0))
        if amount <= 0:
            raise ValueError
    except (ValueError, TypeError):
        messages.error(request, 'Укажите корректную сумму')
        return redirect('main:expenses_list')

    date_str = request.POST.get('date', '')
    try:
        from datetime import date as date_cls
        date_val = date_cls.fromisoformat(date_str)
    except Exception:
        date_val = timezone.now().date()

    Expense.objects.create(
        owner=owner,
        category=request.POST.get('category', 'other'),
        amount=amount,
        description=request.POST.get('description', '').strip(),
        date=date_val,
    )
    messages.success(request, 'Расход добавлен')
    return redirect(f"{request.META.get('HTTP_REFERER', '/expenses/')}")


@login_required
def delete_expense(request, expense_id):
    from apps.main.models import Expense
    owner = get_tenant_owner(request.user)
    expense = get_object_or_404(Expense, id=expense_id, owner=owner)
    if request.method == 'POST':
        expense.delete()
        messages.success(request, 'Расход удалён')
    return redirect('main:expenses_list')

# ─── Файлы к заказу ─────────────────────────────────────────────────────────

@login_required
def upload_order_attachment(request, order_id):
    from apps.rental.models import RentalOrder, OrderAttachment
    owner = get_tenant_owner(request.user)
    order = get_object_or_404(RentalOrder, id=order_id, client__owner=owner)

    if request.method != 'POST':
        return redirect('main:order_view', order_id=order_id)

    f = request.FILES.get('file')
    if not f:
        messages.error(request, 'Выберите файл')
        return redirect('main:order_view', order_id=order_id)

    if f.size > 20 * 1024 * 1024:
        messages.error(request, 'Файл слишком большой (максимум 20 МБ)')
        return redirect('main:order_view', order_id=order_id)

    OrderAttachment.objects.create(
        order=order,
        file=f,
        name=f.name,
        uploaded_by=request.user,
    )
    return redirect('main:order_view', order_id=order_id)


@login_required
def delete_order_attachment(request, attachment_id):
    from apps.rental.models import OrderAttachment
    owner = get_tenant_owner(request.user)
    att = get_object_or_404(OrderAttachment, id=attachment_id, order__client__owner=owner)
    order_id = att.order_id
    if request.method == 'POST':
        att.file.delete(save=False)
        att.delete()
    return redirect('main:order_view', order_id=order_id)


# ─── Excel экспорт ──────────────────────────────────────────────────────────

def _make_xlsx_response(filename):
    from django.http import HttpResponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _style_header(ws, headers, widths=None):
    """Записывает заголовки с форматированием."""
    import openpyxl.styles as st
    hdr_font  = st.Font(bold=True, color='FFFFFF', size=11)
    hdr_fill  = st.PatternFill('solid', fgColor='4F46E5')
    hdr_align = st.Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.alignment = hdr_align
    ws.row_dimensions[1].height = 24
    if widths:
        from openpyxl.utils import get_column_letter
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w


@login_required
def export_clients_xlsx(request):
    import openpyxl
    from django.utils import timezone as tz

    owner = get_tenant_owner(request.user)
    clients = (Client.objects
               .filter(owner=owner)
               .prefetch_related('phones', 'rental_orders')
               .order_by('last_name', 'first_name'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Клиенты'

    headers = ['#', 'Фамилия', 'Имя', 'Отчество', 'Телефон', 'Заказов',
               'Активных', 'Долг (сом)', 'Зарегистрирован']
    widths  = [5, 18, 15, 15, 18, 10, 10, 14, 18]
    _style_header(ws, headers, widths)

    for row_num, client in enumerate(clients, 2):
        phone = client.phones.first()
        debt  = client.get_debt()
        total_orders  = client.rental_orders.count()
        active_orders = client.rental_orders.filter(status='open').count()
        ws.append([
            row_num - 1,
            client.last_name,
            client.first_name,
            client.middle_name,
            phone.phone_number if phone else '',
            total_orders,
            active_orders,
            float(debt),
            client.created_at.strftime('%d.%m.%Y'),
        ])

    response = _make_xlsx_response(f'clients_{tz.now().strftime("%Y%m%d")}.xlsx')
    wb.save(response)
    return response


@login_required
def export_orders_xlsx(request):
    import openpyxl
    import re as _re
    from django.utils import timezone as tz

    owner = get_tenant_owner(request.user)
    status_filter = request.GET.get('status', '')
    orders_qs = RentalOrder.objects.filter(client__owner=owner).select_related('client')
    if status_filter in ('open', 'closed'):
        orders_qs = orders_qs.filter(status=status_filter)
    orders_qs = orders_qs.order_by('-created_at')

    def _paid_for_order(client_obj, order_id_int):
        paid = 0.0
        for p in client_obj.payments.all():
            notes = p.notes or ''
            if 'Распределение:' in notes:
                for line in notes.split('\n'):
                    m = _re.search(rf'Заказ\s*#{order_id_int}\s*:\s*(\d+(?:[\.,]\d+)?)\s*сом', line)
                    if m:
                        paid += float(m.group(1).replace(',', '.'))
                        break
            elif f'Заказ #{order_id_int}' in notes or f'#{order_id_int}' in notes:
                paid += float(p.amount)
        return paid

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Заказы'

    headers = ['#', 'Код заказа', 'Клиент', 'Статус', 'Создан',
               'Сумма (сом)', 'Оплачено (сом)', 'Долг (сом)', 'Доставка']
    widths  = [5, 14, 24, 10, 16, 14, 14, 14, 8]
    _style_header(ws, headers, widths)

    for row_num, order in enumerate(orders_qs, 2):
        total = order.get_current_total()
        paid  = _paid_for_order(order.client, order.id)
        debt  = max(0, float(total) - float(paid))
        ws.append([
            row_num - 1,
            order.order_code,
            order.client.get_full_name(),
            'Открыт' if order.status == 'open' else 'Закрыт',
            order.created_at.strftime('%d.%m.%Y'),
            float(total),
            float(paid),
            debt,
            'Да' if order.has_delivery else 'Нет',
        ])

    response = _make_xlsx_response(f'orders_{tz.now().strftime("%Y%m%d")}.xlsx')
    wb.save(response)
    return response


@login_required
def export_payments_xlsx(request):
    import openpyxl
    from django.utils import timezone as tz

    owner = get_tenant_owner(request.user)
    payments_qs = (Payment.objects
                   .filter(client__owner=owner)
                   .select_related('client')
                   .order_by('-payment_date'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Оплаты'

    headers = ['#', 'Дата', 'Клиент', 'Примечание', 'Сумма (сом)', 'Способ оплаты']
    widths  = [5, 14, 24, 14, 14, 16]
    _style_header(ws, headers, widths)

    method_labels = {'cash': 'Наличные', 'card': 'Карта', 'transfer': 'Перевод'}

    for row_num, payment in enumerate(payments_qs, 2):
        ws.append([
            row_num - 1,
            payment.payment_date.strftime('%d.%m.%Y') if payment.payment_date else '',
            payment.client.get_full_name(),
            payment.notes[:30] if payment.notes else '',
            float(payment.amount),
            method_labels.get(payment.payment_method, payment.payment_method),
        ])

    response = _make_xlsx_response(f'payments_{tz.now().strftime("%Y%m%d")}.xlsx')
    wb.save(response)
    return response


# ══════════════════════════════════════════════════════════════════════════════
# КЛИЕНТСКИЙ ПОРТАЛ — управление заявками
# ══════════════════════════════════════════════════════════════════════════════

@login_required
def bookings_list(request):
    """Список заявок на бронирование. Для создателя — с меню директоров."""
    from apps.main.models import BookingRequest
    from apps.clients.models import Client
    from apps.inventory.models import Product
    from apps.rental.models import RentalOrder, Payment
    from django.db.models import Sum, Count, Q

    owner = get_tenant_owner(request.user)
    is_creator = request.user.is_staff

    directors = []
    selected_director = None

    if is_creator:
        # Список директоров с числом заявок
        director_users = User.objects.filter(is_superuser=True, is_staff=False, is_active=True)
        for d in director_users:
            pending = BookingRequest.objects.filter(client__owner=d, status='pending').count()
            total = BookingRequest.objects.filter(client__owner=d).count()
            clients_count = Client.objects.filter(owner=d).count()
            products_count = Product.objects.filter(owner=d).count()
            revenue = Payment.objects.filter(client__owner=d).aggregate(t=Sum('amount'))['t'] or 0
            open_orders = RentalOrder.objects.filter(client__owner=d, status='open').count()
            directors.append({
                'user': d,
                'pending': pending,
                'total': total,
                'clients': clients_count,
                'products': products_count,
                'revenue': int(revenue),
                'open_orders': open_orders,
            })

        # Выбранный директор
        dir_id = request.GET.get('director', '')
        if dir_id:
            try:
                selected_director = User.objects.get(id=int(dir_id))
                owner = selected_director
            except (User.DoesNotExist, ValueError):
                pass

        if selected_director:
            bookings = BookingRequest.objects.filter(client__owner=selected_director)
        else:
            # Показать все заявки от всех директоров
            bookings = BookingRequest.objects.all()
    else:
        bookings = BookingRequest.objects.filter(client__owner=owner)

    bookings = bookings.select_related('client', 'product').order_by('-created_at')

    status_filter = request.GET.get('status', '')
    if status_filter:
        bookings = bookings.filter(status=status_filter)

    page_obj, page_query = paginate(request, bookings, per_page=20)

    return render(request, 'portal/bookings_manage.html', {
        'bookings': page_obj,
        'page_obj': page_obj,
        'page_query': page_query,
        'status_filter': status_filter,
        'is_creator': is_creator,
        'directors': directors,
        'selected_director': selected_director,
    })


@login_required
def booking_approve(request, booking_id):
    """Одобрить заявку — создать реальный заказ."""
    from apps.main.models import BookingRequest
    from apps.rental.models import RentalOrder, OrderItem
    from decimal import Decimal
    import datetime

    booking = get_object_or_404(BookingRequest, id=booking_id)
    if request.method == 'POST':
        booking.status = 'approved'
        booking.admin_comment = request.POST.get('comment', '').strip()
        booking.save()

        # --- Создаём реальный заказ ---
        product = booking.product
        client = booking.client
        qty = min(booking.quantity, product.quantity_available)

        order = None
        if qty > 0:
            order = RentalOrder.objects.create(
                client=client,
                status='open',
                notes=f'Создан из заявки на бронирование #{booking.id}',
            )

            issued = timezone.make_aware(
                datetime.datetime.combine(booking.start_date, datetime.time(9, 0))
            )
            end_dt = timezone.make_aware(
                datetime.datetime.combine(booking.end_date, datetime.time(18, 0))
            )
            days = max((booking.end_date - booking.start_date).days, 1)
            cost = Decimal(str(product.price_per_day)) * days * qty

            OrderItem.objects.create(
                order=order,
                product=product,
                quantity_taken=qty,
                quantity_remaining=qty,
                issued_date=issued,
                planned_return_date=end_dt,
                rental_days=days,
                rental_hours=0,
                price_per_day=product.price_per_day,
                price_per_hour=product.price_per_hour,
                original_total_cost=cost,
                current_total_cost=cost,
            )

            product.update_available_quantity()

            messages.success(request, f'Заявка #{booking.id} одобрена. Создан заказ #{order.order_number} ({order.order_code})')
        else:
            messages.warning(request, f'Заявка #{booking.id} одобрена, но товар недоступен — заказ не создан')

        # --- Email клиенту ---
        try:
            from apps.main.email_utils import _send
            if client.email:
                body = (
                    f'Здравствуйте, {client.get_full_name()}!\n\n'
                    f'Ваша заявка на бронирование одобрена.\n'
                    f'Товар: {product.name}\n'
                    f'Период: {booking.start_date.strftime("%d.%m.%Y")} — {booking.end_date.strftime("%d.%m.%Y")}\n'
                    f'Количество: {qty} шт.\n\n'
                    f'Пожалуйста, приходите в назначенную дату.\n\nС уважением,\nСлужба аренды'
                )
                _send(f'Заявка одобрена — {product.name}', body, client.email)
        except Exception:
            pass

    return redirect('main:bookings_list')


@login_required
def booking_reject(request, booking_id):
    """Отклонить заявку на бронирование."""
    from apps.main.models import BookingRequest
    booking = get_object_or_404(BookingRequest, id=booking_id)
    if request.method == 'POST':
        booking.status = 'rejected'
        booking.admin_comment = request.POST.get('comment', '').strip()
        booking.save()

        try:
            from apps.main.email_utils import _send
            if booking.client.email:
                body = (
                    f'Здравствуйте, {booking.client.get_full_name()}!\n\n'
                    f'К сожалению, ваша заявка на бронирование отклонена.\n'
                    f'Товар: {booking.product.name}\n'
                )
                if booking.admin_comment:
                    body += f'Причина: {booking.admin_comment}\n\n'
                body += 'Если у вас есть вопросы, свяжитесь с нами.\n\nС уважением,\nСлужба аренды'
                _send(f'Заявка отклонена — {booking.product.name}', body, booking.client.email)
        except Exception:
            pass

        messages.success(request, f'Заявка #{booking.id} отклонена')
    return redirect('main:bookings_list')


@login_required
def send_portal_link(request, client_id):
    """Сгенерировать токен портала и отправить ссылку клиенту."""
    from apps.main.models import ClientPortalToken
    client = get_object_or_404(Client, id=client_id)

    token_obj, created = ClientPortalToken.objects.get_or_create(client=client)

    portal_url = request.build_absolute_uri(f'/portal/{token_obj.token}/')

    sent_via = []
    # Email
    if client.email:
        try:
            from apps.main.email_utils import _send
            body = (
                f'Здравствуйте, {client.get_full_name()}!\n\n'
                f'Ваша персональная ссылка для бронирования инструментов:\n'
                f'{portal_url}\n\n'
                f'По этой ссылке вы можете:\n'
                f'- Посмотреть каталог доступных инструментов\n'
                f'- Забронировать нужные инструменты\n'
                f'- Отслеживать статус ваших заявок и заказов\n\n'
                f'С уважением,\nСлужба аренды'
            )
            if _send('Ваша ссылка для бронирования', body, client.email):
                sent_via.append('Email')
        except Exception:
            pass

    # Telegram
    if client.telegram_id:
        try:
            from apps.main.telegram_bot_complete import send_telegram_message
            text = (
                f'Здравствуйте, {client.get_full_name()}!\n\n'
                f'Ваша ссылка для бронирования:\n{portal_url}'
            )
            if send_telegram_message(client.telegram_id, text, parse_mode=None):
                sent_via.append('Telegram')
        except Exception:
            pass

    if sent_via:
        messages.success(request, f'Ссылка на портал отправлена: {", ".join(sent_via)}')
    else:
        messages.info(request, f'Ссылка: {portal_url} (у клиента нет email/telegram для автоотправки)')

    return redirect('main:client_detail', client_id=client.id)


# === СКРЫТЫЙ АУДИТ-ЦЕНТР (только суперюзер) ===

import shutil
from pathlib import Path

_BACKUP_DIR = Path(__file__).resolve().parent.parent.parent / 'backups'
_DB_PATH    = Path(__file__).resolve().parent.parent.parent / 'db.sqlite3'


def _superuser_required(view_func):
    from functools import wraps
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated or not request.user.is_superuser:
            from django.http import HttpResponseForbidden
            return HttpResponseForbidden('403 Forbidden')
        return view_func(request, *args, **kwargs)
    return wrapper


@_superuser_required
def xsec_audit(request):
    from apps.main.models import RequestLog, ActivityLog
    user_filter  = request.GET.get('u', '')
    path_filter  = request.GET.get('p', '')
    event_filter = request.GET.get('ev', '')
    date_from    = request.GET.get('df', '')
    date_to      = request.GET.get('dt', '')
    tab          = request.GET.get('tab', 'requests')

    rlogs = RequestLog.objects.select_related('user').order_by('-created_at')
    if user_filter:  rlogs = rlogs.filter(username__icontains=user_filter)
    if path_filter:  rlogs = rlogs.filter(path__icontains=path_filter)
    if event_filter == 'login':  rlogs = rlogs.filter(event='login')
    elif event_filter == 'logout': rlogs = rlogs.filter(event='logout')
    elif event_filter == 'close':  rlogs = rlogs.filter(event='page_close')
    elif event_filter == 'errors': rlogs = rlogs.filter(status_code__gte=400)
    if date_from: rlogs = rlogs.filter(created_at__date__gte=date_from)
    if date_to:   rlogs = rlogs.filter(created_at__date__lte=date_to)
    page_obj, page_query = paginate(request, rlogs, per_page=100)

    alogs = ActivityLog.objects.select_related('user').order_by('-created_at')
    if user_filter: alogs = alogs.filter(user__username__icontains=user_filter)
    if date_from: alogs = alogs.filter(created_at__date__gte=date_from)
    if date_to:   alogs = alogs.filter(created_at__date__lte=date_to)
    alogs_page, alogs_query = paginate(request, alogs, per_page=100)

    _BACKUP_DIR.mkdir(exist_ok=True)
    backups = sorted(_BACKUP_DIR.glob('db_backup_*.sqlite3'),
                     key=lambda f: f.stat().st_mtime, reverse=True)
    backup_list = [{'name': bf.name,
                    'size_kb': round(bf.stat().st_size / 1024, 1),
                    'mtime': datetime.fromtimestamp(bf.stat().st_mtime).strftime('%d.%m.%Y %H:%M:%S')}
                   for bf in backups]

    total_requests = RequestLog.objects.count()
    total_logins   = RequestLog.objects.filter(event='login').count()
    total_errors   = RequestLog.objects.filter(status_code__gte=400).count()
    active_users   = RequestLog.objects.filter(
        created_at__gte=timezone.now() - timedelta(hours=24)
    ).values('username').distinct().count()

    return render(request, 'main/xsec_audit.html', {
        'tab': tab, 'page_obj': page_obj, 'page_query': page_query,
        'alogs': alogs_page, 'alogs_query': alogs_query,
        'backups': backup_list,
        'total_requests': total_requests, 'total_logins': total_logins,
        'total_errors': total_errors, 'active_users': active_users,
        'user_filter': user_filter, 'path_filter': path_filter,
        'event_filter': event_filter, 'date_from': date_from, 'date_to': date_to,
    })


@_superuser_required
def xsec_backup_create(request):
    if request.method != 'POST': return redirect('/xsec-audit/?tab=backups')
    try:
        _BACKUP_DIR.mkdir(exist_ok=True)
        ts   = timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M%S')
        dest = _BACKUP_DIR / ('db_backup_' + ts + '.sqlite3')
        shutil.copy2(_DB_PATH, dest)
        all_b = sorted(_BACKUP_DIR.glob('db_backup_*.sqlite3'), key=lambda f: f.stat().st_mtime)
        for old in all_b[:-30]: old.unlink(missing_ok=True)
        messages.success(request, 'Бэкап создан: ' + dest.name)
    except Exception as e:
        messages.error(request, 'Ошибка: ' + str(e))
    return redirect('/xsec-audit/?tab=backups')


@_superuser_required
def xsec_backup_restore(request, backup_name):
    if request.method != 'POST': return redirect('/xsec-audit/?tab=backups')
    try:
        src = _BACKUP_DIR / backup_name
        if not src.exists() or not backup_name.startswith('db_backup_'):
            messages.error(request, 'Бэкап не найден'); return redirect('/xsec-audit/?tab=backups')
        ts = timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M%S')
        shutil.copy2(_DB_PATH, _BACKUP_DIR / ('db_backup_before_restore_' + ts + '.sqlite3'))
        shutil.copy2(src, _DB_PATH)
        messages.success(request, 'База восстановлена из ' + backup_name)
    except Exception as e:
        messages.error(request, 'Ошибка: ' + str(e))
    return redirect('/xsec-audit/?tab=backups')


@_superuser_required
def xsec_backup_download(request, backup_name):
    src = _BACKUP_DIR / backup_name
    if not src.exists() or not backup_name.startswith('db_backup_'): raise Http404
    return FileResponse(open(src, 'rb'), as_attachment=True, filename=backup_name)


@_superuser_required
def xsec_backup_delete(request, backup_name):
    if request.method != 'POST': return redirect('/xsec-audit/?tab=backups')
    try:
        src = _BACKUP_DIR / backup_name
        if src.exists() and backup_name.startswith('db_backup_'):
            src.unlink(); messages.success(request, 'Удалён: ' + backup_name)
    except Exception as e:
        messages.error(request, 'Ошибка: ' + str(e))
    return redirect('/xsec-audit/?tab=backups')


@login_required
def xsec_beacon(request):
    if request.method == 'POST':
        try:
            import json as _json
            from apps.main.models import RequestLog
            body = _json.loads(request.body or b'{}')
            ip = (request.META.get('HTTP_X_FORWARDED_FOR','') or request.META.get('REMOTE_ADDR',''))
            if ',' in ip: ip = ip.split(',')[0].strip()
            RequestLog.objects.create(
                user=request.user, username=request.user.username, ip=ip or None,
                method='BEACON', path=body.get('path','/')[:500],
                status_code=200, user_agent=request.META.get('HTTP_USER_AGENT','')[:500],
                event='page_close',
            )
        except Exception: pass
    from django.http import HttpResponse
    return HttpResponse('', status=204)
